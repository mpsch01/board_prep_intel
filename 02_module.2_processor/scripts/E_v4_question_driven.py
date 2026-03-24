"""
E_v4_question_driven.py
=======================
Question-driven session matching — cleanest version.

Key improvements over E_v2/v3:
  1. Reads from ABFM_ITE_Master_v2.xlsx (authoritative question bank)
  2. Merges AI_ClinicalFocus from AI_Tagged archive at runtime (join on QID)
  3. Uses refs_2025_extracted.csv directly for 2025 ref lookup
     (no Explanation-based parsing needed — Master v2 Reference column has all refs)
  4. Uses ITE_Reference_Tiers_Clean.csv (existing, 1,065 refs)
  5. Output: session_hy_inserts_v6.json

Ref strategy:
  - For ALL questions: use Reference column from Master v2 (pipe-separated if multiple)
  - Fallback: parse Explanation column (catches any remaining inline refs)
  - This gives full 2020-2025 coverage since Master v2 now has 200/200 2025 refs

Output: aafp_integration/02_working/session_hy_inserts_v6.json

Author: Pipeline auto-generated 2026-03-03
"""

import json, csv, re, os
from pathlib import Path
from openpyxl import load_workbook
from difflib import SequenceMatcher
from collections import defaultdict

# ── Paths ──────────────────────────────────────────────────────────────────────
SCRIPT_DIR   = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent.parent
KW_JSON   = PROJECT_ROOT / "key_data_files" / "session_keyword_library.json"
MASTER    = PROJECT_ROOT / "key_data_files" / "ABFM_ITE_Master_v2.xlsx"           # TODO: not yet migrated
AI_TAGGED = PROJECT_ROOT / "key_data_files" / "ABFM_ITE_AI_Tagged_excel.xlsx"     # TODO: not yet migrated
TIER_CSV  = PROJECT_ROOT / "key_data_files" / "ITE_Reference_Tiers_Clean.csv"     # TODO: not yet migrated
CW_CSV    = PROJECT_ROOT / "key_data_files" / "session_cluster_crosswalk.csv"     # output of 01_build_crosswalk.py
OUT_JSON  = PROJECT_ROOT / "key_data_files" / "session_hy_inserts_v6.json"
VAL_CSV   = SCRIPT_DIR.parent / "outputs" / "E_v4_validation.csv"

# ── Tuning ─────────────────────────────────────────────────────────────────────
MAX_Q_PER_SESSION = 5
MIN_KW_SCORE      = 0.10
MIN_KW_HITS       = 2
MATCH_THRESHOLD   = 0.70

# ── Stopwords ──────────────────────────────────────────────────────────────────
STOPWORDS = {
    'the','a','an','of','in','to','and','or','is','are','was','were','be',
    'been','being','have','has','had','do','does','did','will','would','could',
    'should','may','might','can','this','that','these','those','with','for',
    'from','not','but','at','by','as','it','its','he','she','they','their',
    'which','who','what','when','where','how','one','two','three','also',
    'most','more','other','following','patient','patients','which','year',
    'old','male','female','history','presents','exam','test','use','used',
    'using','associated','include','including','such','based','than','both',
    'after','before','during','without','within','per','type','level','rate',
    'new','well','age','used','first','initial','given','known','seen',
    'risk','blood','pressure','treatment','diagnosis','management','clinical',
}

def tokenize_bigrams(text):
    tokens = re.findall(r'[a-zA-Z]{3,}', (text or '').lower())
    filtered = [t for t in tokens if t not in STOPWORDS]
    unigrams = set(filtered)
    bigrams  = {f'{filtered[i]} {filtered[i+1]}' for i in range(len(filtered)-1)}
    return unigrams | bigrams

# ── Load keyword library ───────────────────────────────────────────────────────
def load_keyword_library(path):
    with open(path, encoding='utf-8') as f:
        raw = json.load(f)
    sessions = {}
    for sess_id, data in raw.items():
        terms = {kw['term'].lower(): kw['composite'] for kw in data.get('keywords', [])}
        sessions[sess_id] = {
            'session_name':  data.get('session_name', ''),
            'terms':         terms,
            'max_composite': max(terms.values()) if terms else 1.0
        }
    return sessions

# ── Load tier DB ───────────────────────────────────────────────────────────────
def load_tier_db(path):
    rows = []
    with open(path, newline='', encoding='utf-8') as f:
        for row in csv.DictReader(f):
            norm = re.sub(r'[^\w\s]', ' ', row['CleanRef'].lower())
            norm = re.sub(r'\s+', ' ', norm).strip()
            rows.append({'CleanRef': row['CleanRef'], 'Tier': row['Tier'], 'norm': norm})
    return rows

def fuzzy_match_ref(raw, tier_db):
    norm_raw = re.sub(r'[^\w\s]', ' ', raw.lower())
    norm_raw = re.sub(r'\s+', ' ', norm_raw).strip()
    best_score, best = 0.0, None
    for row in tier_db:
        s = SequenceMatcher(None, norm_raw[:200], row['norm'][:200]).ratio()
        if s > best_score:
            best_score, best = s, row
    if best_score >= MATCH_THRESHOLD and best:
        return best['CleanRef'], best['Tier'], round(best_score, 3)
    return raw.strip(), 'Unmatched', round(best_score, 3)

# ── Parse refs from Reference column (pipe-separated) ─────────────────────────
CITE_HINT = re.compile(r'\b(19|20)\d{2}\b')

def parse_refs_from_reference_col(ref_str):
    """Split pipe-separated refs from Master v2 Reference column."""
    if not ref_str or str(ref_str).strip() in ('', 'nan', 'None'):
        return []
    parts = [r.strip() for r in str(ref_str).split('|')]
    return [p for p in parts if len(p) > 15 and CITE_HINT.search(p)]

# ── Parse refs from Explanation column (2020-2024 fallback) ──────────────────
REF_INLINE  = re.compile(r'Ref:\s*(.+)', re.DOTALL | re.IGNORECASE)
NUM_SPLIT   = re.compile(r'\s+\d+\)\s+')

def parse_refs_from_explanation(expl):
    if not expl:
        return []
    refs = []
    m = REF_INLINE.search(expl)
    if m:
        ref_block = m.group(1).strip()
        parts = NUM_SPLIT.split(ref_block)
        for p in parts:
            p = p.strip().rstrip('0123456789').strip()
            if len(p) > 20 and CITE_HINT.search(p):
                refs.append(p)
    return refs

# ── Load AI_ClinicalFocus from archived xlsx ───────────────────────────────────
def load_ai_focus(path):
    """Returns dict: QuestionID → AI_ClinicalFocus string."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_raw = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() if h else '' for h in rows_raw[0]]
    lookup = {}
    qid_col    = headers.index('Question ID')   if 'Question ID'    in headers else None
    focus_col  = headers.index('AI_ClinicalFocus') if 'AI_ClinicalFocus' in headers else None
    subcat_col = headers.index('AI_Subcategory')   if 'AI_Subcategory'   in headers else None
    for row in rows_raw[1:]:
        qid   = str(row[qid_col] or '').strip()   if qid_col is not None else ''
        focus = str(row[focus_col] or '').strip()  if focus_col is not None else ''
        subcat= str(row[subcat_col] or '').strip() if subcat_col is not None else ''
        if qid:
            lookup[qid] = f'{focus} {subcat}'.strip()
    wb.close()
    print(f'  AI focus loaded: {len(lookup)} QIDs')
    return lookup

# ── Load master questions ──────────────────────────────────────────────────────
def load_questions(master_path, ai_focus_lookup):
    wb = load_workbook(master_path, read_only=True, data_only=True)
    ws = wb.active
    rows_raw = list(ws.iter_rows(values_only=True))
    headers  = [str(h).strip() if h else '' for h in rows_raw[0]]
    questions = []
    for row in rows_raw[1:]:
        d = dict(zip(headers, row))
        qid = str(d.get('QuestionID', '') or '')
        questions.append({
            'qid':        qid,
            'year':       int(d.get('ExamYear', 0) or 0),
            'stem':       str(d.get('QuestionStem', '') or ''),
            'explanation':str(d.get('Explanation', '') or ''),
            'reference':  str(d.get('Reference', '') or ''),
            'subcategory':str(d.get('Subcategory', '') or ''),
            'ai_focus':   ai_focus_lookup.get(qid, ''),
        })
    wb.close()
    return questions

# ── Score one question against one session ─────────────────────────────────────
def score_question_vs_session(q_tokens, session):
    terms = session['terms']
    max_c = session['max_composite']
    score = 0.0
    hits  = 0
    for tok in q_tokens:
        if tok in terms:
            score += terms[tok] / max_c
            hits  += 1
    return round(score, 4), hits

# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    print('Loading keyword library...')
    sessions = load_keyword_library(KW_JSON)
    print(f'  {len(sessions)} sessions')

    print('Loading tier DB...')
    tier_db = load_tier_db(TIER_CSV)
    print(f'  {len(tier_db)} refs')

    print('Loading crosswalk...')
    with open(CW_CSV, newline='', encoding='utf-8') as f:
        cw = {r['session_number']: r for r in csv.DictReader(f)}

    print('Loading AI_ClinicalFocus from archive...')
    ai_focus = load_ai_focus(AI_TAGGED)

    print('Loading master questions...')
    questions = load_questions(MASTER, ai_focus)
    print(f'  {len(questions)} questions')

    # Coverage check
    refs_filled = sum(1 for q in questions if parse_refs_from_reference_col(q['reference']))
    print(f'  Questions with Reference: {refs_filled}/1200')
    by_year = {}
    for q in questions:
        yr = q['year']
        if yr not in by_year:
            by_year[yr] = {'total':0,'ref':0}
        by_year[yr]['total'] += 1
        if parse_refs_from_reference_col(q['reference']):
            by_year[yr]['ref'] += 1
    for yr in sorted(by_year):
        d = by_year[yr]
        print(f'    {yr}: {d["ref"]}/{d["total"]} refs')

    # ── Score every question against every session ─────────────────────────────
    print('\nScoring questions vs sessions...')
    session_matches = defaultdict(list)

    for q in questions:
        # Token set: stem + AI_ClinicalFocus + Subcategory
        q_text   = f"{q['stem']} {q['ai_focus']} {q['subcategory']}"
        q_tokens = tokenize_bigrams(q_text)

        best_score, best_hits, best_sess = 0.0, 0, None
        second_score, second_sess = 0.0, None

        for sess_id, sess_data in sessions.items():
            sc, hits = score_question_vs_session(q_tokens, sess_data)
            if sc > best_score:
                second_score, second_sess = best_score, best_sess
                best_score, best_hits, best_sess = sc, hits, sess_id
            elif sc > second_score:
                second_score, second_sess = sc, sess_id

        if best_sess and best_score >= MIN_KW_SCORE and best_hits >= MIN_KW_HITS:
            session_matches[best_sess].append((best_score, best_hits, q))

        if (second_sess and second_score >= MIN_KW_SCORE
                and best_hits >= MIN_KW_HITS
                and second_score >= best_score * 0.80):
            session_matches[second_sess].append((second_score, best_hits, q))

    # ── Per session: top N questions, collect refs ─────────────────────────────
    print('Building session inserts...')
    output   = {}
    val_rows = []
    tier_order = {'Must-Read': 0, 'Core': 1, 'Supplementary': 2, 'Unmatched': 3}

    for sess_id in sorted(sessions.keys()):
        matches = session_matches.get(sess_id, [])
        seen_qids = set()
        ranked = []
        for sc, hits, q in sorted(matches, key=lambda x: -x[0]):
            if q['qid'] not in seen_qids:
                seen_qids.add(q['qid'])
                ranked.append((sc, hits, q))

        top_q = ranked[:MAX_Q_PER_SESSION]

        ref_pool = {}
        for sc, hits, q in top_q:
            # Primary: Reference column (pipe-separated, already clean)
            raw_refs = parse_refs_from_reference_col(q['reference'])
            # Fallback: Explanation column parsing
            if not raw_refs:
                raw_refs = parse_refs_from_explanation(q['explanation'])

            for raw in raw_refs:
                matched, tier, mscore = fuzzy_match_ref(raw, tier_db)
                if matched not in ref_pool:
                    ref_pool[matched] = {'tier': tier, 'match_score': mscore, 'qids': []}
                ref_pool[matched]['qids'].append(q['qid'])

        sorted_refs = sorted(
            ref_pool.items(),
            key=lambda x: (tier_order.get(x[1]['tier'], 9), -x[1]['match_score'])
        )

        output[sess_id] = {
            'session_id':          sess_id,
            'session_title':       sessions[sess_id]['session_name'],
            'question_count':      len(top_q),
            'questions': [
                {
                    'qid':          q['qid'],
                    'year':         q['year'],
                    'focus':        q['ai_focus'] or q['subcategory'],
                    'stem_preview': q['stem'][:150],
                    'kw_score':     sc,
                    'kw_hits':      hits,
                }
                for sc, hits, q in top_q
            ],
            'refs': [
                {
                    'citation':    ref,
                    'tier':        info['tier'],
                    'match_score': info['match_score'],
                    'cited_by':    info['qids'],
                }
                for ref, info in sorted_refs
            ],
            'must_read_count':    sum(1 for _, i in sorted_refs if i['tier'] == 'Must-Read'),
            'core_count':         sum(1 for _, i in sorted_refs if i['tier'] == 'Core'),
            'supplementary_count':sum(1 for _, i in sorted_refs if i['tier'] == 'Supplementary'),
        }

        for sc, hits, q in top_q:
            val_rows.append({
                'session_id':   sess_id,
                'session_title':sessions[sess_id]['session_name'],
                'qid':          q['qid'],
                'year':         q['year'],
                'kw_score':     sc,
                'kw_hits':      hits,
                'ai_focus':     q['ai_focus'] or q['subcategory'],
            })

    # ── Write outputs ──────────────────────────────────────────────────────────
    print(f'\nWriting {OUT_JSON}...')
    with open(OUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(output, f, indent=2)

    os.makedirs(os.path.dirname(VAL_CSV), exist_ok=True)
    with open(VAL_CSV, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=['session_id','session_title','qid','year','kw_score','kw_hits','ai_focus'])
        w.writeheader()
        w.writerows(val_rows)
    print(f'Validation CSV: {VAL_CSV}')

    # ── Summary ────────────────────────────────────────────────────────────────
    total_q    = sum(v['question_count'] for v in output.values())
    total_refs = sum(len(v['refs']) for v in output.values())
    total_mr   = sum(v['must_read_count'] for v in output.values())
    total_core = sum(v['core_count'] for v in output.values())
    no_q       = sum(1 for v in output.values() if v['question_count'] == 0)
    no_mr      = sum(1 for v in output.values() if v['must_read_count'] == 0)

    print('\n=== SUMMARY ===')
    print(f'  Sessions with questions : {len(output)-no_q}/{len(output)}')
    print(f'  Sessions with must-read : {len(output)-no_mr}/{len(output)}')
    print(f'  Total questions placed  : {total_q}')
    print(f'  Total refs placed       : {total_refs}')
    print(f'  Must-Read refs          : {total_mr}')
    print(f'  Core refs               : {total_core}')
    print(f'  Sessions w/ 0 questions : {no_q}')

    print('\nSpot-check top 3 sessions:')
    for sid in list(sorted(output.keys()))[:3]:
        v = output[sid]
        print(f'  Session {sid} [{v["session_title"]}]: {v["question_count"]} Qs, {len(v["refs"])} refs')
        for q in v['questions'][:2]:
            print(f'    [{q["year"]}] {q["focus"][:60]} (score={q["kw_score"]})')

    print('\nDone.')

if __name__ == '__main__':
    main()
