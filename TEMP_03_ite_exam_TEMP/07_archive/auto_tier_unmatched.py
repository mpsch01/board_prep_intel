"""
auto_tier_unmatched.py
======================
Assigns tiers to the 281 unmatched refs using source-type heuristics:
  - AFP article (Am Fam Physician) → Core
  - JAMA / NEJM / Lancet / Annals review → Core
  - Major guideline (AHA, ACC, USPSTF, ADA, etc.) → Core
  - Textbook chapter → Supplementary
  - Case report / letter / everything else → Supplementary

Must-Read requires human judgment — not auto-assigned.

Then merges with existing tier DB to create ITE_Reference_Tiers_Expanded.csv
"""

import csv, re
from pathlib import Path
from collections import Counter

ALIGNED_CSV = Path(r"C:\Users\mpsch\Desktop\claude_knowledge\board_prep\ite_exam\02_working\ref_tier_aligned.csv")
TIER_CSV    = Path(r"C:\Users\mpsch\Desktop\claude_knowledge\board_prep\ite_refs\03_database\ITE_Reference_Tiers_Final.csv")
OUT_CSV     = Path(r"C:\Users\mpsch\Desktop\claude_knowledge\board_prep\ite_refs\03_database\ITE_Reference_Tiers_Expanded.csv")
MASTER_XLSX = Path(r"C:\Users\mpsch\Desktop\claude_knowledge\board_prep\ite_exam\03_database\ABFM_ITE_Master_v2.xlsx")

def infer_source_type(ref: str) -> str:
    r = ref.lower()
    if re.search(r'am\s*fam\s*physician', r):           return 'AFP'
    if re.search(r'\bjama\b', r):                        return 'JAMA'
    if re.search(r'n\s*engl\s*j\s*med', r):             return 'NEJM'
    if re.search(r'\blancet\b', r):                      return 'Lancet'
    if re.search(r'ann\s*intern\s*med', r):              return 'Annals'
    if re.search(r'\bcirculation\b', r):                 return 'Circulation'
    if re.search(r'\bpediatrics\b', r):                  return 'Pediatrics'
    if re.search(r'obstet\s*gynecol', r):                return 'ObGyn'
    if re.search(r'bmj|brit\s*med\s*j', r):             return 'BMJ'
    if re.search(r'thyroid|chest|stroke|diabetes care', r): return 'SpecialtyJournal'
    if re.search(r'guideline|recommendation|statement|task force|advisory', r): return 'Guideline'
    if re.search(r'harrison|goldman|williams|sabiston|mandell', r): return 'Textbook'
    if re.search(r'dsm|icd|uptodate', r):               return 'Reference'
    return 'Other'

def infer_tier(ref: str, source_type: str) -> str:
    """Heuristic tier assignment — never assigns Must-Read (human judgment required)"""
    if source_type in ('AFP', 'JAMA', 'NEJM', 'Lancet', 'Annals'):
        return 'Core'
    if source_type == 'Guideline':
        return 'Core'
    if source_type in ('Circulation', 'Pediatrics', 'ObGyn', 'BMJ', 'SpecialtyJournal'):
        return 'Core'
    if source_type in ('Textbook', 'Reference'):
        return 'Supplementary'
    return 'Supplementary'

# Load aligned CSV to find unmatched
aligned = []
with open(ALIGNED_CSV, encoding="utf-8") as f:
    for row in csv.DictReader(f):
        aligned.append(row)

unmatched = [r for r in aligned if not r['Tier']]
matched   = [r for r in aligned if r['Tier']]
print(f"Total aligned: {len(aligned)}")
print(f"Previously matched: {len(matched)}")
print(f"Unmatched to auto-tier: {len(unmatched)}")

# Count ref frequency from master
import openpyxl
wb = openpyxl.load_workbook(MASTER_XLSX, read_only=True)
ws = wb["Sheet1"]
header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
col_ref = header.index("Reference")
col_year = header.index("ExamYear")

ref_freq = Counter()
ref_years = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    raw = row[col_ref]
    yr = str(row[col_year]) if row[col_year] else "?"
    if raw:
        for part in str(raw).split("|"):
            part = part.strip()
            if part:
                ref_freq[part] += 1
                ref_years.setdefault(part, set()).add(yr)

# Auto-tier unmatched refs
auto_tiered = []
tier_counts = Counter()
source_counts = Counter()

for r in unmatched:
    ref = r['Reference']
    src = infer_source_type(ref)
    tier = infer_tier(ref, src)
    freq = ref_freq.get(ref, 1)
    years = ','.join(sorted(ref_years.get(ref, set())))
    tier_counts[tier] += 1
    source_counts[src] += 1
    auto_tiered.append({
        'CleanRef': ref,
        'CitationCount': freq,
        'UniqueYears': len(ref_years.get(ref, set())),
        'SourceType': src,
        'Categories': '',
        'BlueprintCategories': '',
        'Tier': tier,
        'AutoAssigned': 'Yes',
    })

print(f"\nAuto-tier results:")
for tier, cnt in tier_counts.most_common():
    print(f"  {tier}: {cnt}")
print(f"\nSource type breakdown:")
for src, cnt in source_counts.most_common():
    print(f"  {src}: {cnt}")

# Load existing tier DB
existing = []
with open(TIER_CSV, encoding="utf-8") as f:
    for row in csv.DictReader(f):
        existing.append({
            'CleanRef': row['CleanRef'],
            'CitationCount': row['CitationCount'],
            'UniqueYears': row['UniqueYears'],
            'SourceType': row['SourceType'],
            'Categories': row['Categories'],
            'BlueprintCategories': row['BlueprintCategories'],
            'Tier': row['Tier'],
            'AutoAssigned': 'No',
        })

# Combine
combined = existing + auto_tiered
print(f"\nExpanded tier DB: {len(existing)} original + {len(auto_tiered)} auto = {len(combined)} total")

# Save
fieldnames = ['CleanRef','CitationCount','UniqueYears','SourceType','Categories','BlueprintCategories','Tier','AutoAssigned']
with open(OUT_CSV, "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(combined)
print(f"Saved: {OUT_CSV}")
print(f"\nTier DB now covers {len(combined)} unique references.")
print("Must-Read candidates flagged for manual review:")
# Show AFP refs cited 3+ times that got auto-tiered as Core (potential Must-Read upgrades)
must_read_candidates = [r for r in auto_tiered 
                        if r['Tier'] == 'Core' and int(r.get('CitationCount',0)) >= 3]
must_read_candidates.sort(key=lambda x: -int(x.get('CitationCount',0)))
for r in must_read_candidates[:20]:
    print(f"  [{r['CitationCount']}x] {r['CleanRef'][:85]}")
