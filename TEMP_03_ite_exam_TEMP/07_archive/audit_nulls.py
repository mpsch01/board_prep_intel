"""
audit_nulls.py — Check actual null state across all years in master Excel
"""
import openpyxl
from pathlib import Path
from collections import defaultdict

XLSX = Path(r"C:\Users\mpsch\Desktop\claude_knowledge\board_prep\ite_exam\03_database\ABFM_ITE_Master_v2.xlsx")

wb = openpyxl.load_workbook(XLSX, read_only=True)
ws = wb["Sheet1"]

header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
col_qid  = header.index("QuestionID")
col_year = header.index("ExamYear")
col_ref  = header.index("Reference")
col_exp  = header.index("Explanation")

stats = defaultdict(lambda: {"total":0,"ref_null":0,"exp_null":0})

for row in ws.iter_rows(min_row=2, values_only=True):
    qid  = row[col_qid]
    year = str(row[col_year]) if row[col_year] else "?"
    ref  = row[col_ref]
    exp  = row[col_exp]
    stats[year]["total"] += 1
    if not ref or str(ref).strip() in ("", "None"):
        stats[year]["ref_null"] += 1
    if not exp or str(exp).strip() in ("", "None"):
        stats[year]["exp_null"] += 1

print(f"\n{'Year':<8} {'Total':>7} {'Ref Null':>10} {'Exp Null':>10} {'Ref%':>8}")
print("-" * 48)
grand_total = grand_ref_null = grand_exp_null = 0
for year in sorted(stats.keys()):
    s = stats[year]
    pct = 100 * s["ref_null"] / s["total"] if s["total"] else 0
    print(f"{year:<8} {s['total']:>7} {s['ref_null']:>10} {s['exp_null']:>10} {pct:>7.1f}%")
    grand_total    += s["total"]
    grand_ref_null += s["ref_null"]
    grand_exp_null += s["exp_null"]

print("-" * 48)
pct = 100 * grand_ref_null / grand_total if grand_total else 0
print(f"{'TOTAL':<8} {grand_total:>7} {grand_ref_null:>10} {grand_exp_null:>10} {pct:>7.1f}%")
print()
