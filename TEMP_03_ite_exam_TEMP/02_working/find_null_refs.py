"""
find_null_refs.py — identify the 7 null reference rows and cross-match with critique text
"""
import openpyxl, json
from pathlib import Path

XLSX     = Path(r"C:\Users\mpsch\Desktop\claude_knowledge\board_prep\ite_exam\03_database\ABFM_ITE_Master_v2.xlsx")
JSON_25  = Path(r"C:\Users\mpsch\Desktop\claude_knowledge\board_prep\ite_exam\02_working\refs_2025_extracted.json")

refs_25 = json.loads(JSON_25.read_text(encoding="utf-8"))

wb = openpyxl.load_workbook(XLSX)
ws = wb["Sheet1"]

header = [cell.value for cell in ws[1]]
col_qid  = header.index("QuestionID") + 1
col_year = header.index("ExamYear")   + 1
col_ref  = header.index("Reference")  + 1

print("Null reference rows:")
print(f"{'QID':<15} {'Year':>6}  {'In 2025 JSON':>14}  {'Ref from JSON'}")
print("-"*80)

for row in ws.iter_rows(min_row=2, values_only=False):
    qid  = row[col_qid  - 1].value
    year = str(row[col_year - 1].value)
    ref  = row[col_ref  - 1].value
    if ref and str(ref).strip() not in ("", "None"):
        continue
    # null ref found
    from_json = refs_25.get(qid, [])
    print(f"{qid:<15} {year:>6}  {str(bool(from_json)):>14}  {' | '.join(from_json[:2]) if from_json else 'NOT IN JSON'}")
