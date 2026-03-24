"""
extract_2025_refs.py
====================
Parses 2025_ITE_Critique.txt → extracts References block for each question
(sequentially mapped Q2025-001 through Q2025-200)
Then backfills ABFM_ITE_Master_v2.xlsx Reference column (col O)
for any 2025 row where Reference is currently null/empty.

Outputs:
  - refs_2025_extracted.json     (audit trail: QID → [ref list])
  - refs_2025_backfill_log.csv   (which cells were updated)
  - ABFM_ITE_Master_v2.xlsx      (updated in-place, col O backfilled)
"""

import re
import json
import csv
import openpyxl
from pathlib import Path

# ── Paths ──────────────────────────────────────────────────────────────────
BASE      = Path(r"C:\Users\mpsch\Desktop\claude_knowledge\board_prep\ite_exam")
CRITIQUE  = BASE / "02_working" / "2025_ITE_Critique.txt"
XLSX      = BASE / "03_database" / "ABFM_ITE_Master_v2.xlsx"
OUT_DIR   = BASE / "02_working"

JSON_OUT  = OUT_DIR / "refs_2025_extracted.json"
LOG_OUT   = OUT_DIR / "refs_2025_backfill_log.csv"

# ── Step 1: Parse critique file ────────────────────────────────────────────
text = CRITIQUE.read_text(encoding="utf-8", errors="replace")

# Split on "ANSWER:" boundaries — each chunk = one question's block
# First chunk before the first ANSWER is the header → skip it
chunks = re.split(r'(?=ANSWER:\s*[A-E])', text)
chunks = [c.strip() for c in chunks if re.match(r'ANSWER:\s*[A-E]', c.strip())]

print(f"Total ANSWER blocks found: {len(chunks)}")

def extract_refs(chunk: str) -> list[str]:
    """Pull lines after 'References' header until next blank-line block."""
    lines = chunk.splitlines()
    in_refs = False
    refs = []
    for line in lines:
        stripped = line.strip()
        if re.match(r'^References\s*$', stripped, re.IGNORECASE):
            in_refs = True
            continue
        if in_refs:
            # Stop if we hit another ANSWER block or blank line after refs started
            if re.match(r'^ANSWER:\s*[A-E]', stripped):
                break
            if stripped:
                refs.append(stripped)
    return refs

# Build QID → refs dict
qid_refs = {}
for i, chunk in enumerate(chunks):
    qid = f"Q2025-{i+1:03d}"
    refs = extract_refs(chunk)
    qid_refs[qid] = refs

# Save JSON audit
JSON_OUT.write_text(json.dumps(qid_refs, indent=2), encoding="utf-8")
print(f"Saved: {JSON_OUT}")

# Summary stats
total_qs   = len(qid_refs)
with_refs  = sum(1 for r in qid_refs.values() if r)
empty_refs = total_qs - with_refs
total_ref_strings = sum(len(r) for r in qid_refs.values())
print(f"Questions parsed: {total_qs}")
print(f"  With refs: {with_refs}  |  Empty: {empty_refs}")
print(f"  Total reference strings: {total_ref_strings}")

# ── Step 2: Backfill Excel ─────────────────────────────────────────────────
wb = openpyxl.load_workbook(XLSX)
ws = wb["Sheet1"]

# Find column indices (1-based)
header = [cell.value for cell in ws[1]]
col_qid  = header.index("QuestionID") + 1   # col A = 1
col_year = header.index("ExamYear")    + 1
col_ref  = header.index("Reference")   + 1   # col O

print(f"\nColumn mapping: QuestionID={col_qid}, ExamYear={col_year}, Reference={col_ref}")

log_rows = [["QuestionID", "ExamYear", "Action", "ReferenceCount", "ReferencesAdded"]]
updated = 0
skipped_existing = 0
skipped_no_refs  = 0

for row in ws.iter_rows(min_row=2, values_only=False):
    qid  = row[col_qid  - 1].value
    year = row[col_year - 1].value
    ref_cell = row[col_ref - 1]

    if str(year) != "2025":
        continue

    existing = ref_cell.value
    new_refs = qid_refs.get(qid, [])

    if existing and str(existing).strip():
        # Already has reference — skip
        skipped_existing += 1
        log_rows.append([qid, year, "SKIPPED_HAS_REF", len(new_refs), ""])
        continue

    if not new_refs:
        skipped_no_refs += 1
        log_rows.append([qid, year, "SKIPPED_NO_REFS_FOUND", 0, ""])
        continue

    # Join multiple refs with " | "
    combined = " | ".join(new_refs)
    ref_cell.value = combined
    updated += 1
    log_rows.append([qid, year, "UPDATED", len(new_refs), combined])

print(f"\nBackfill results:")
print(f"  Updated:          {updated}")
print(f"  Skipped (exists): {skipped_existing}")
print(f"  Skipped (no ref): {skipped_no_refs}")

# Save workbook
wb.save(XLSX)
print(f"\nSaved updated Excel: {XLSX}")

# Save log
with open(LOG_OUT, "w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerows(log_rows)
print(f"Saved log: {LOG_OUT}")
print("\nDone.")
