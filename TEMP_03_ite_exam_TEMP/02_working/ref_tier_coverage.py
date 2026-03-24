"""
ref_tier_coverage.py
Check what % of references in the master Excel have tier assignments in the tier DB
Identify high-frequency untiered refs that should be prioritized for manual assignment
"""
import openpyxl, csv, re
from pathlib import Path
from collections import Counter

XLSX     = Path(r"C:\Users\mpsch\Desktop\claude_knowledge\board_prep\ite_exam\03_database\ABFM_ITE_Master_v2.xlsx")
TIER_CSV = Path(r"C:\Users\mpsch\Desktop\claude_knowledge\board_prep\ite_refs\03_database\ITE_Reference_Tiers_Final.csv")
OUT_DIR  = Path(r"C:\Users\mpsch\Desktop\claude_knowledge\board_prep\ite_exam\02_working")

# Load tier DB — build lookup by normalized title/author fragment
tier_rows = []
with open(TIER_CSV, encoding="utf-8") as f:
    reader = csv.DictReader(f)
    tier_rows = list(reader)

print(f"Tier DB rows: {len(tier_rows)}")
print(f"Tier DB columns: {list(tier_rows[0].keys()) if tier_rows else 'empty'}")

# Build set of tiered reference strings (lowercased, stripped)
tiered_refs = set()
for r in tier_rows:
    # Try all text columns for matching
    for col in ["Reference", "Full_Citation", "Title", "Citation", "reference", "full_citation"]:
        val = r.get(col, "")
        if val and str(val).strip():
            tiered_refs.add(str(val).strip().lower()[:80])  # first 80 chars as key

print(f"Unique tiered ref keys: {len(tiered_refs)}")

# Extract all refs from master Excel
wb = openpyxl.load_workbook(XLSX, read_only=True)
ws = wb["Sheet1"]
header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
col_ref  = header.index("Reference")
col_year = header.index("ExamYear")

all_ref_strings = []
year_ref_counts = {}

for row in ws.iter_rows(min_row=2, values_only=True):
    year = str(row[col_year]) if row[col_year] else "?"
    ref  = row[col_ref]
    if not ref or str(ref).strip() in ("", "None"):
        continue
    # Split pipe-delimited refs
    parts = [r.strip() for r in str(ref).split("|") if r.strip()]
    all_ref_strings.extend(parts)
    year_ref_counts.setdefault(year, Counter())
    for p in parts:
        year_ref_counts[year][p] += 1

print(f"\nTotal reference strings across all questions: {len(all_ref_strings)}")
print(f"Unique references: {len(set(all_ref_strings))}")

# Check tier coverage
in_tier = 0
not_in_tier = []
ref_counter = Counter(all_ref_strings)

for ref_str, count in ref_counter.most_common():
    key = ref_str.strip().lower()[:80]
    # Fuzzy: check if any tiered ref starts with same 40 chars
    matched = any(key[:40] in t for t in tiered_refs)
    if matched:
        in_tier += count
    else:
        not_in_tier.append((ref_str, count))

total_ref_instances = sum(ref_counter.values())
print(f"\nTier coverage: {in_tier}/{total_ref_instances} = {100*in_tier/total_ref_instances:.1f}%")
print(f"Untiered unique refs: {len(not_in_tier)}")

# Show top 30 untiered by frequency
print("\nTop 30 untiered references by frequency:")
print(f"{'Count':>6}  Reference")
print("-"*90)
for ref, cnt in sorted(not_in_tier, key=lambda x: -x[1])[:30]:
    print(f"{cnt:>6}  {ref[:85]}")

# Save full untiered list
out_path = OUT_DIR / "untiered_refs_priority.csv"
with open(out_path, "w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerow(["Frequency", "Reference"])
    for ref, cnt in sorted(not_in_tier, key=lambda x: -x[1]):
        writer.writerow([cnt, ref])
print(f"\nSaved: {out_path}")
