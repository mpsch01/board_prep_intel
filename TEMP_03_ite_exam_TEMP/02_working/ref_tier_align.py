"""
ref_tier_align.py
=================
Aligns master Excel references against ITE_Reference_Tiers_Final.csv
using multi-strategy fuzzy matching:
  1. First-author last name + year + journal abbreviation
  2. Title keyword overlap (Jaccard similarity)
  3. First 60 chars after normalization

Outputs:
  - ref_tier_aligned.csv     (each unique master ref → matched tier or None)
  - ref_tier_coverage_v2.csv (summary stats)
  - untiered_refs_v2.csv     (unmatched refs for manual tier assignment)
"""

import openpyxl, csv, re
from pathlib import Path
from collections import Counter

XLSX     = Path(r"C:\Users\mpsch\Desktop\claude_knowledge\board_prep\ite_exam\03_database\ABFM_ITE_Master_v2.xlsx")
TIER_CSV = Path(r"C:\Users\mpsch\Desktop\claude_knowledge\board_prep\ite_refs\03_database\ITE_Reference_Tiers_Final.csv")
OUT_DIR  = Path(r"C:\Users\mpsch\Desktop\claude_knowledge\board_prep\ite_exam\02_working")

def normalize(s):
    """Lowercase, remove punctuation, collapse spaces"""
    s = str(s).lower()
    s = re.sub(r'[^\w\s]', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()

def extract_year(s):
    m = re.search(r'\b(19|20)\d{2}\b', s)
    return m.group(0) if m else ""

def first_author(s):
    """Extract first author last name"""
    # Try 'Last, F.' format first, then 'Last F' 
    m = re.match(r'([A-Z][a-z]+)[,\s]', s)
    return m.group(1).lower() if m else ""

def journal_abbrev(s):
    """Look for Am Fam Physician, JAMA, NEJM, Lancet, etc."""
    patterns = {
        'afp': r'am\s*fam\s*physician',
        'jama': r'\bjama\b',
        'nejm': r'n\s*engl\s*j\s*med',
        'lancet': r'\blancet\b',
        'bmj': r'\bbmj\b',
        'annals': r'ann\s*intern\s*med',
        'circulation': r'\bcirculation\b',
        'pediatrics': r'\bpediatrics\b',
        'obstet': r'obstet\s*gynecol',
        'thyroid': r'\bthyroid\b',
        'chest': r'\bchest\b',
    }
    s_low = s.lower()
    for abbr, pat in patterns.items():
        if re.search(pat, s_low):
            return abbr
    return ""

def jaccard(a, b, n=3):
    """n-gram Jaccard similarity"""
    a_set = set(a[i:i+n] for i in range(len(a)-n+1))
    b_set = set(b[i:i+n] for i in range(len(b)-n+1))
    if not a_set or not b_set:
        return 0
    return len(a_set & b_set) / len(a_set | b_set)

# Load tier DB
tier_rows = []
with open(TIER_CSV, encoding="utf-8") as f:
    for row in csv.DictReader(f):
        tier_rows.append({
            'raw': row['CleanRef'],
            'norm': normalize(row['CleanRef']),
            'year': extract_year(row['CleanRef']),
            'author': first_author(row['CleanRef']),
            'journal': journal_abbrev(row['CleanRef']),
            'tier': row['Tier'],
            'source_type': row.get('SourceType',''),
        })
print(f"Tier DB loaded: {len(tier_rows)} refs")

# Load master Excel refs
wb = openpyxl.load_workbook(XLSX, read_only=True)
ws = wb["Sheet1"]
header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
col_ref  = header.index("Reference")
col_year = header.index("ExamYear")

ref_year_map = {}  # ref_string → set of exam years
for row in ws.iter_rows(min_row=2, values_only=True):
    exam_year = str(row[col_year]) if row[col_year] else "?"
    ref = row[col_ref]
    if not ref or str(ref).strip() in ("", "None"):
        continue
    for part in str(ref).split("|"):
        part = part.strip()
        if part:
            ref_year_map.setdefault(part, set()).add(exam_year)

unique_refs = list(ref_year_map.keys())
print(f"Unique master refs: {len(unique_refs)}")

# Match each master ref against tier DB
results = []
match_counts = Counter()

for ref_str in unique_refs:
    norm_ref = normalize(ref_str)
    year_r   = extract_year(ref_str)
    auth_r   = first_author(ref_str)
    jour_r   = journal_abbrev(ref_str)
    
    best_score = 0
    best_match = None
    best_strategy = None
    
    for t in tier_rows:
        score = 0
        strategy = []
        
        # Strategy 1: exact first 60 chars after normalization
        if norm_ref[:60] == t['norm'][:60] and len(norm_ref) > 30:
            score = 100; strategy = ['exact60']
        else:
            # Strategy 2: author + year + journal
            if auth_r and auth_r == t['author']:
                score += 35; strategy.append('author')
            if year_r and year_r == t['year']:
                score += 25; strategy.append('year')
            if jour_r and jour_r == t['journal']:
                score += 20; strategy.append('journal')
            # Strategy 3: Jaccard on normalized string
            j = jaccard(norm_ref, t['norm'])
            score += j * 40
            if j > 0.4:
                strategy.append(f'jaccard={j:.2f}')
        
        if score > best_score:
            best_score = score
            best_match = t
            best_strategy = strategy
    
    tier = best_match['tier'] if best_match and best_score >= 55 else None
    match_counts[f"score>={55}" if tier else "unmatched"] += 1
    
    results.append({
        'Reference': ref_str,
        'ExamYears': ','.join(sorted(ref_year_map[ref_str])),
        'MatchScore': round(best_score, 1),
        'Strategy': '|'.join(best_strategy) if best_strategy else '',
        'MatchedRef': best_match['raw'] if best_match else '',
        'Tier': tier or '',
        'SourceType': best_match['source_type'] if best_match and tier else '',
    })

# Save aligned CSV
aligned_path = OUT_DIR / "ref_tier_aligned.csv"
with open(aligned_path, "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=['Reference','ExamYears','MatchScore','Strategy','MatchedRef','Tier','SourceType'])
    writer.writeheader()
    writer.writerows(results)
print(f"Saved: {aligned_path}")

# Coverage summary
matched   = [r for r in results if r['Tier']]
unmatched = [r for r in results if not r['Tier']]
print(f"\nCoverage: {len(matched)}/{len(results)} = {100*len(matched)/len(results):.1f}%")
print(f"Unmatched: {len(unmatched)}")

# Tier breakdown
tier_counts = Counter(r['Tier'] for r in matched)
print("\nTier breakdown (matched refs):")
for tier, cnt in tier_counts.most_common():
    print(f"  {tier}: {cnt}")

# Top unmatched by frequency
ref_freq = Counter()
for row in ws.iter_rows(min_row=2, values_only=True):
    raw = row[col_ref]
    if raw:
        for part in str(raw).split("|"):
            part = part.strip()
            if part:
                ref_freq[part] += 1

unmatched_path = OUT_DIR / "untiered_refs_v2.csv"
with open(unmatched_path, "w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerow(["Reference", "ExamYears", "BestScore", "SuggestedTier"])
    for r in sorted(unmatched, key=lambda x: -x['MatchScore'])[:200]:
        writer.writerow([r['Reference'], r['ExamYears'], r['MatchScore'], ''])
print(f"Saved untiered list: {unmatched_path}")
