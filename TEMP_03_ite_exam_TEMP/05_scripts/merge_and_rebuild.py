#!/usr/bin/env python3
"""
merge_and_rebuild.py

Merges all new derived fields into the enriched master database, then
rebuilds every year-specific Excel file (ITE_2020.xlsx … ITE_2025.xlsx).

New fields added to the master:
  BodySystem            — extracted (PDF gold standard) or ML-predicted
  BodySystem_Source     — "Extracted" | "Predicted"
  BodySystem_Confidence — decision-function confidence (1.0 for gold)
  ClinicalCategory      — extracted from 2022-2023 PDFs; blank otherwise
  Subcategory           — question-action type (Pharmacology, Screening, etc.)
  Subcat_Source         — high / medium / low / fallback
  Reference             — clinical literature citation (from QA source)

Blueprint / Clinical category strategy (per user):
  - 2024-2025: BlueprintCategory populated natively (leave as-is)
  - 2022-2023: ClinicalCategory added as separate column
  - 2020-2021: Neither — both columns blank for these years

Output files:
  board_prep/ite_exam/03_database/
    ABFM_ITE_Master_v2.xlsx          — full 1200-row master with all fields
    raw_files/ABFM_ITE_Master_v2.csv — same, as CSV backup

  board_prep/ite_exam/04_outputs/by_year/
    ITE_2020.xlsx … ITE_2025.xlsx     — one formatted workbook per year
    raw_files/ITE_2020_raw.csv … ITE_2025_raw.csv
"""

import pandas as pd
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment,
                              Border, Side, numbers)
from openpyxl.utils import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────────────────
ITE_BASE  = Path(__file__).resolve().parents[1]
PIPE_BASE = ITE_BASE.parent / "ite_pipeline/03_database"

SRC_ENRICHED  = ITE_BASE / "03_database/ABFM_ITE_Enriched.xlsx"
SRC_BODY_SYS  = ITE_BASE / "03_database/body_system_full.csv"
SRC_CLIN_CAT  = PIPE_BASE / "clinical_category_labels.csv"
SRC_SUBCAT    = ITE_BASE / "03_database/subcategory_labels.csv"

# Reference sources
REFS_BASE     = ITE_BASE.parent / "ite_refs/02_working"
SRC_REF_PAIRS = REFS_BASE / "question_ref_pairs.csv"   # primary: 1085 questions mapped

OUT_DB       = ITE_BASE / "03_database/ABFM_ITE_Master_v2.xlsx"
OUT_DB_RAW   = ITE_BASE / "03_database/raw_files/ABFM_ITE_Master_v2.csv"
OUT_YEAR_DIR = ITE_BASE / "04_outputs/by_year"
RAW_YEAR_DIR = ITE_BASE / "04_outputs/by_year/raw_files"

for d in [OUT_DB.parent / "raw_files", OUT_YEAR_DIR, RAW_YEAR_DIR]:
    d.mkdir(parents=True, exist_ok=True)


# ═══════════════════════════════════════════════════════════════════════════════
# 1. LOAD & MERGE
# ═══════════════════════════════════════════════════════════════════════════════

print("Loading base enriched master...")
df = pd.read_excel(SRC_ENRICHED)
df = df.rename(columns={"Question ID": "QuestionID"})
# Drop original empty columns that will be re-populated with richer data
df = df.drop(columns=["Subcategory", "References"], errors="ignore")
print(f"  {len(df)} rows  |  cols: {list(df.columns)}")

# ── Body System ───────────────────────────────────────────────────────────────
print("\nMerging BodySystem labels...")
df_bs = pd.read_csv(SRC_BODY_SYS)[
    ["QuestionID", "BodySystem", "BodySystem_Source", "BodySystem_Confidence"]
]
df = df.merge(df_bs, on="QuestionID", how="left")
print(f"  BodySystem populated: {df['BodySystem'].notna().sum()}")

# ── Clinical Category (2022-2023 only) ────────────────────────────────────────
print("Merging ClinicalCategory labels (2022-2023)...")
df_cc = pd.read_csv(SRC_CLIN_CAT)

# Clinical category labels use per-year QuestionNum (1-200) not master IDs
# Need positional join for 2022-2023
df["WithinYearPos"] = df.groupby("ExamYear").cumcount() + 1
df_cc_pos = df_cc.rename(columns={"Year": "ExamYear", "QuestionNum": "WithinYearPos"})

df = df.merge(
    df_cc_pos[["ExamYear", "WithinYearPos", "ClinicalCategory"]],
    on=["ExamYear", "WithinYearPos"],
    how="left"
)
print(f"  ClinicalCategory populated: {df['ClinicalCategory'].notna().sum()} "
      f"(expected ~393 for 2022-2023)")

# ── Subcategory ───────────────────────────────────────────────────────────────
print("Merging Subcategory...")
df_sub = pd.read_csv(SRC_SUBCAT)[["QuestionID", "Subcategory", "Subcat_Source"]]
df = df.merge(df_sub, on="QuestionID", how="left")
print(f"  Subcategory populated: {df['Subcategory'].notna().sum()}")

# ── References (two-source merge) ────────────────────────────────────────────
# Primary:   question_ref_pairs.csv — direct QuestionID match, 1085 questions
# Secondary: subcategory_labels.csv (QA-extracted) — fills 2024 gaps
print("Merging References (primary: question_ref_pairs; secondary: QA-extracted)...")

df_ref_qa = pd.read_csv(SRC_SUBCAT)[["QuestionID", "Reference"]].rename(
    columns={"Reference": "Ref_QA"}
)

df_ref_pairs_raw = pd.read_csv(SRC_REF_PAIRS)
df_ref_pairs = (
    df_ref_pairs_raw.groupby("QuestionID")["RefMatched"]
    .apply(lambda refs: " | ".join(
        r for r in refs if str(r).strip() not in ("", "nan")
    ))
    .reset_index()
    .rename(columns={"RefMatched": "Ref_Pairs"})
)

df = df.merge(df_ref_pairs, on="QuestionID", how="left")
df = df.merge(df_ref_qa,    on="QuestionID", how="left")

# Prefer pairs; fall back to QA-extracted
def best_ref(row):
    p = str(row.get("Ref_Pairs", "")).strip()
    q = str(row.get("Ref_QA",    "")).strip()
    if p and p != "nan": return p
    if q and q != "nan": return q
    return ""

df["Reference"] = df.apply(best_ref, axis=1)
df = df.drop(columns=["Ref_Pairs", "Ref_QA"], errors="ignore")

ref_cov = df.groupby("ExamYear")["Reference"].apply(lambda x: (x != "").sum())
print("  Reference coverage by year:")
for yr, n in ref_cov.items():
    print(f"    {yr}: {n}/200")


# ═══════════════════════════════════════════════════════════════════════════════
# 2. COLUMN ORDER FOR MASTER
# ═══════════════════════════════════════════════════════════════════════════════

MASTER_COLS = [
    # Identity
    "QuestionID", "ExamYear",
    # Question content
    "QuestionStem", "CorrectAnswer",
    # Body system classification (gold-standard PDF extraction + ML; replaces legacy PrimaryCategory)
    "BodySystem", "BodySystem_Source", "BodySystem_Confidence",
    # Year-specific clinical categorization
    "ClinicalCategory",                          # 2022-2023 only (extracted from score reports)
    "BlueprintCategory",                         # 2024-2025 only (gold)
    "BlueprintCategory_Predicted", "Blueprint_Confidence", "Blueprint_Source",
    # Question-type
    "Subcategory", "Subcat_Source",
    # Clinical reference
    "Reference",
    # Explanation
    "Explanation",
    # Admin
    "ScoringStatus", "YearLastUsed",
]

# Keep only columns that exist
MASTER_COLS = [c for c in MASTER_COLS if c in df.columns]
df_master = df[MASTER_COLS].copy()

print(f"\nMaster ready: {df_master.shape}  |  final columns: {len(MASTER_COLS)}")

# Quick stats
print("\n=== BodySystem distribution ===")
print(df_master["BodySystem"].value_counts().head(10).to_string())
print("\n=== Subcategory distribution ===")
print(df_master["Subcategory"].value_counts().to_string())
print("\n=== Reference coverage by year ===")
ref_cov = df_master.groupby("ExamYear")["Reference"].apply(lambda x: (x != '').sum())
print(ref_cov.to_string())

# Save master
df_master.to_excel(OUT_DB, index=False)
df_master.to_csv(OUT_DB_RAW, index=False)
print(f"\n  Saved master: {OUT_DB.name}")
print(f"  Saved CSV:    {OUT_DB_RAW.name}")


# ═══════════════════════════════════════════════════════════════════════════════
# 3. REBUILD YEAR-BY-YEAR EXCEL FILES
# ═══════════════════════════════════════════════════════════════════════════════

# Columns shown in each year file
YEAR_COLS = [
    "QuestionID", "ExamYear", "QuestionStem", "CorrectAnswer",
    "BodySystem", "ClinicalCategory", "BlueprintCategory",
    "Subcategory", "Reference", "Explanation",
    "BlueprintCategory_Predicted", "Blueprint_Confidence",
    "BodySystem_Source", "BodySystem_Confidence", "Subcat_Source",
]
YEAR_COLS = [c for c in YEAR_COLS if c in df_master.columns]

# Excel styling constants
HDR_FILL   = PatternFill("solid", fgColor="1F3864")   # dark navy
ALT_FILL   = PatternFill("solid", fgColor="EBF0FA")   # light blue
WHT_FILL   = PatternFill("solid", fgColor="FFFFFF")
HDR_FONT   = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
BODY_FONT  = Font(size=9, name="Calibri")
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
TOP_ALIGN  = Alignment(vertical="top")

THIN = Side(style="thin", color="CCCCCC")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COL_WIDTHS = {
    "QuestionID":                  14,
    "ExamYear":                    10,
    "QuestionStem":                55,
    "CorrectAnswer":               14,
    "PrimaryCategory":             22,
    "BodySystem":                  26,
    "BodySystem_Source":           18,
    "BodySystem_Confidence":       20,
    "ClinicalCategory":            24,
    "BlueprintCategory":           24,
    "Subcategory":                 20,
    "Subcat_Source":               14,
    "Reference":                   50,
    "Explanation":                 55,
    "BlueprintCategory_Predicted": 26,
    "Blueprint_Confidence":        20,
}

def make_year_workbook(year_df: pd.DataFrame, year: int) -> Workbook:
    wb = Workbook()

    # ── Summary sheet ────────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    summaries = [
        ("Body System",        "BodySystem"),
        ("Blueprint Category", "BlueprintCategory"),
        ("Clinical Category",  "ClinicalCategory"),
        ("Subcategory",        "Subcategory"),
        ("Primary Category",   "PrimaryCategory"),
    ]

    col_offset = 1
    for label, field in summaries:
        if field not in year_df.columns:
            continue
        counts = year_df[field].value_counts(dropna=True)
        if counts.empty:
            continue

        # Header row
        h_label = ws_sum.cell(row=1, column=col_offset, value=label)
        h_count = ws_sum.cell(row=1, column=col_offset + 1, value="Count")
        h_pct   = ws_sum.cell(row=1, column=col_offset + 2, value="%")
        for cell in [h_label, h_count, h_pct]:
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = TOP_ALIGN

        total = len(year_df)
        for r, (cat, n) in enumerate(counts.items(), start=2):
            fill = ALT_FILL if r % 2 == 0 else WHT_FILL
            for c, val in zip(
                [col_offset, col_offset + 1, col_offset + 2],
                [cat, n, round(n / total * 100, 1)]
            ):
                cell = ws_sum.cell(row=r, column=c, value=val)
                cell.fill = fill
                cell.font = BODY_FONT
                cell.alignment = TOP_ALIGN

        ws_sum.column_dimensions[get_column_letter(col_offset)].width     = 28
        ws_sum.column_dimensions[get_column_letter(col_offset + 1)].width = 10
        ws_sum.column_dimensions[get_column_letter(col_offset + 2)].width = 8
        col_offset += 4   # gap between tables

    ws_sum.freeze_panes = "A2"

    # ── Questions sheet ──────────────────────────────────────────────────────
    ws_q = wb.create_sheet("Questions")

    display_cols = [c for c in YEAR_COLS if c in year_df.columns]
    year_df_out  = year_df[display_cols].fillna("")

    # Header
    for c_idx, col_name in enumerate(display_cols, start=1):
        cell = ws_q.cell(row=1, column=c_idx, value=col_name)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center",
                                   horizontal="center")
        ws_q.column_dimensions[get_column_letter(c_idx)].width = (
            COL_WIDTHS.get(col_name, 18)
        )

    ws_q.row_dimensions[1].height = 30

    # Data rows
    for r_idx, (_, row) in enumerate(year_df_out.iterrows(), start=2):
        fill = ALT_FILL if r_idx % 2 == 0 else WHT_FILL
        for c_idx, col_name in enumerate(display_cols, start=1):
            val  = row[col_name]
            cell = ws_q.cell(row=r_idx, column=c_idx, value=val)
            cell.fill      = fill
            cell.font      = BODY_FONT
            cell.border    = CELL_BORDER
            cell.alignment = WRAP_ALIGN
        ws_q.row_dimensions[r_idx].height = 80

    ws_q.freeze_panes = "A2"
    ws_q.auto_filter.ref = ws_q.dimensions

    return wb


print("\nBuilding year files...")
years = sorted(df_master["ExamYear"].unique())
for yr in years:
    yr_df = df_master[df_master["ExamYear"] == yr].copy()

    # Raw CSV
    raw_path = RAW_YEAR_DIR / f"ITE_{yr}_raw.csv"
    yr_df.to_csv(raw_path, index=False)

    # Formatted Excel
    xl_path = OUT_YEAR_DIR / f"ITE_{yr}.xlsx"
    wb = make_year_workbook(yr_df, yr)
    wb.save(xl_path)

    # Subcategory breakdown for quick log
    sub_counts = yr_df["Subcategory"].value_counts()
    top3 = ", ".join(f"{k}({v})" for k, v in sub_counts.head(3).items())
    print(f"  {yr}: {len(yr_df)} questions  |  top subcats: {top3}  → {xl_path.name}")

print("\nAll year files rebuilt.")
print("\nDone.")
