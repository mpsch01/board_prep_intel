#!/usr/bin/env python3
"""
split_by_year.py
Splits FINAL_ABFM_ITE_6_Year_Master.xlsx into one formatted Excel file per exam year.
Outputs to: 04_outputs/by_year/
Raw CSVs to: 04_outputs/by_year/raw_files/
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

BASE     = Path(__file__).resolve().parents[1]
SRC      = BASE / "03_database/ABFM_ITE_Enriched.xlsx"
OUT_DIR  = BASE / "04_outputs/by_year"
RAW_DIR  = OUT_DIR / "raw_files"
OUT_DIR.mkdir(parents=True, exist_ok=True)
RAW_DIR.mkdir(parents=True, exist_ok=True)

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")   # dark navy
ALT_FILL     = PatternFill("solid", fgColor="DCE6F1")   # light blue
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT    = Font(name="Arial", size=10)
CENTER       = Alignment(horizontal="center", vertical="top", wrap_text=False)
LEFT_WRAP    = Alignment(horizontal="left", vertical="top", wrap_text=True)
LEFT         = Alignment(horizontal="left", vertical="top")
THIN         = Side(style="thin", color="B8CCE4")
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COL_WIDTHS = {
    "Question ID":      14,
    "ExamYear":         10,
    "QuestionStem":     80,
    "CorrectAnswer":    16,
    "PrimaryCategory":  22,
    "Subcategory":      22,
    "Explanation":      80,
    "References":       30,
    "YearLastUsed":     14,
}

DISPLAY_COLS = [
    "Question ID", "ExamYear", "QuestionStem", "CorrectAnswer",
    "PrimaryCategory", "Subcategory", "Explanation", "References", "YearLastUsed",
    "BlueprintCategory", "ScoringStatus", "BlueprintCategory_Predicted",
    "Blueprint_Confidence", "Blueprint_Source", "Subcategory_Cluster", "Cluster_Keywords"
]

COL_WIDTHS.update({
    "BlueprintCategory":          24,
    "ScoringStatus":              14,
    "BlueprintCategory_Predicted": 26,
    "Blueprint_Confidence":       20,
    "Blueprint_Source":           16,
    "Subcategory_Cluster":        24,
    "Cluster_Keywords":           40,
})

def style_sheet(ws, df):
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = COL_WIDTHS.get(col_name, 18)

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else WHITE_FILL
        for col_idx, (col_name, value) in enumerate(zip(df.columns, row), start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value if pd.notna(value) else ""
            cell.font = BODY_FONT
            cell.fill = fill
            cell.border = BORDER
            if col_name in ("QuestionStem", "Explanation"):
                cell.alignment = LEFT_WRAP
            elif col_name in ("Question ID", "ExamYear", "CorrectAnswer", "YearLastUsed"):
                cell.alignment = CENTER
            else:
                cell.alignment = LEFT

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 22

def add_summary_sheet(wb, df, year):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    title_font  = Font(name="Arial", bold=True, size=14, color="1F4E79")
    label_font  = Font(name="Arial", bold=True, size=10)
    value_font  = Font(name="Arial", size=10)
    sect_fill   = PatternFill("solid", fgColor="1F4E79")
    sect_font   = Font(name="Arial", bold=True, color="FFFFFF", size=10)

    ws["B2"] = f"ITE {year} — Exam Summary"
    ws["B2"].font = title_font
    ws["B3"] = f"Source: FINAL_ABFM_ITE_6_Year_Master.xlsx  |  {len(df)} questions"
    ws["B3"].font = Font(name="Arial", size=9, italic=True, color="595959")

    ws["B5"] = "Total Questions"
    ws["C5"] = len(df)
    ws["B5"].font = label_font
    ws["C5"].font = value_font

    ws["B7"] = "Category Breakdown"
    ws["B7"].font = sect_font
    ws["B7"].fill = sect_fill
    ws["C7"].fill = sect_fill
    ws["D7"].fill = sect_fill

    ws["B8"]  = "Category"
    ws["C8"]  = "Count"
    ws["D8"]  = "% of Exam"
    for cell in [ws["B8"], ws["C8"], ws["D8"]]:
        cell.font = label_font
        cell.fill = PatternFill("solid", fgColor="DCE6F1")
        cell.border = BORDER
        cell.alignment = CENTER

    cats = df["PrimaryCategory"].value_counts().reset_index()
    cats.columns = ["Category", "Count"]

    for i, row in cats.iterrows():
        r = 9 + i
        ws.cell(row=r, column=2).value = row["Category"]
        ws.cell(row=r, column=3).value = row["Count"]
        ws.cell(row=r, column=4).value = f"=C{r}/C5"
        for c in [2, 3, 4]:
            cell = ws.cell(row=r, column=c)
            cell.font = value_font
            cell.border = BORDER
            fill = PatternFill("solid", fgColor="DCE6F1") if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            cell.fill = fill
        ws.cell(row=r, column=4).number_format = "0.0%"
        ws.cell(row=r, column=3).alignment = CENTER
        ws.cell(row=r, column=4).alignment = CENTER

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12


def build_year_file(df_master, year):
    df = df_master[df_master["ExamYear"] == year][DISPLAY_COLS].reset_index(drop=True)

    # Raw CSV
    df.to_csv(RAW_DIR / f"ITE_{year}_raw.csv", index=False)

    # Formatted Excel
    out_path = OUT_DIR / f"ITE_{year}.xlsx"
    df.to_excel(out_path, index=False, sheet_name=f"ITE {year}")

    wb = load_workbook(out_path)
    ws = wb[f"ITE {year}"]
    ws.sheet_view.showGridLines = False
    style_sheet(ws, df)
    add_summary_sheet(wb, df, year)
    wb.move_sheet("Summary", offset=-(len(wb.sheetnames)-1))
    wb.save(out_path)

    print(f"  ✓ ITE_{year}.xlsx  ({len(df)} questions)")
    return out_path


def main():
    print(f"Reading master: {SRC.name}")
    df_master = pd.read_excel(SRC)
    years = sorted(df_master["ExamYear"].unique())
    print(f"Years found: {years}\n")

    built = []
    for year in years:
        built.append(build_year_file(df_master, year))

    print(f"\nDone. {len(built)} files written to:")
    print(f"  {OUT_DIR}")
    print(f"  {RAW_DIR}  (CSVs)")

if __name__ == "__main__":
    main()
