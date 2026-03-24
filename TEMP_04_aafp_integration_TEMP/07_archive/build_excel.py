import pandas as pd
import numpy as np
import json, re, pickle
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

BASE = Path(r'C:\Users\mpsch\Desktop\claude_knowledge\board_prep\aafp_integration\01_source\updated_data_docs')
OUT  = BASE / 'ITE_Board_Analysis_2020_2025.xlsx'

# ── Reload data ───────────────────────────────────────────────────────────────
df  = pd.read_excel(BASE / 'ABFM_ITE_Master_v2.xlsx')
qrp = pd.read_csv(BASE / 'question_ref_pairs.csv')
with open(BASE / 'session_hy_inserts_v6.json') as f:
    v6 = json.load(f)

df['ExamYear'] = df['ExamYear'].astype(int)
years = sorted(df['ExamYear'].unique())

# ── Style helpers ─────────────────────────────────────────────────────────────
NAVY   = '1F3864'
BLUE   = '2E75B6'
LBLUE  = 'D6E4F7'
AMBER  = 'FFF2CC'
GREEN  = 'E2EFDA'
RED_L  = 'FCE4D6'
WHITE  = 'FFFFFF'
GREY   = 'F2F2F2'
DGREY  = 'D9D9D9'

def hdr(ws, row, col, text, bg=NAVY, fg=WHITE, bold=True, sz=11, wrap=False, halign='center'):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name='Arial', bold=bold, color=fg, size=sz)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal=halign, vertical='center', wrap_text=wrap)
    return c

def val(ws, row, col, v, fmt=None, bg=None, bold=False, halign='center', fg='000000'):
    c = ws.cell(row=row, column=col, value=v)
    c.font = Font(name='Arial', size=10, bold=bold, color=fg)
    c.alignment = Alignment(horizontal=halign, vertical='center')
    if bg:
        c.fill = PatternFill('solid', fgColor=bg)
    if fmt:
        c.number_format = fmt
    return c

def thin_border(ws, r1, c1, r2, c2, color='AAAAAA'):
    side = Side(style='thin', color=color)
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            ws.cell(r,c).border = Border(left=side, right=side, top=side, bottom=side)

def col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def section_title(ws, row, col, text, span=8):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name='Arial', bold=True, size=12, color=NAVY)
    c.fill = PatternFill('solid', fgColor=LBLUE)
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 18

wb = Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1: OVERVIEW DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = 'Overview'
ws.sheet_view.showGridLines = False

# Title block
ws.merge_cells('A1:L1')
c = ws['A1']
c.value = 'ABFM ITE BOARD ANALYSIS  |  2020 – 2025  |  1,200 Questions'
c.font = Font(name='Arial', bold=True, size=16, color=WHITE)
c.fill = PatternFill('solid', fgColor=NAVY)
c.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 32

ws.merge_cells('A2:L2')
c = ws['A2']
c.value = 'Family Medicine Residency Program  ·  ITE High-Yield Intelligence Report  ·  Generated 2026-03-03'
c.font = Font(name='Arial', size=10, color='444444')
c.fill = PatternFill('solid', fgColor=LBLUE)
c.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[2].height = 16

# KPI boxes — row 4
kpis = [
    ('Total Questions', '1,200', '6 exam years'),
    ('Body Systems', '19', 'distinct categories'),
    ('Subcategories', '11', 'question types'),
    ('High-Yield Combos', '20', 'top pockets analyzed'),
    ('References Tracked', '2,069', 'Q→Ref pairs'),
    ('AAFP Sessions', '48', 'fully matched'),
]
ws.row_dimensions[3].height = 8
ws.row_dimensions[4].height = 44
ws.row_dimensions[5].height = 16
ws.row_dimensions[6].height = 8

for i, (label, num, sub) in enumerate(kpis):
    col = i * 2 + 1
    ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+1)
    c = ws.cell(row=4, column=col, value=num)
    c.font = Font(name='Arial', bold=True, size=22, color=BLUE)
    c.fill = PatternFill('solid', fgColor=GREY)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col+1)
    c2 = ws.cell(row=5, column=col, value=label)
    c2.font = Font(name='Arial', bold=True, size=9, color=NAVY)
    c2.fill = PatternFill('solid', fgColor=GREY)
    c2.alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col+1)
    c3 = ws.cell(row=6, column=col, value=sub)
    c3.font = Font(name='Arial', size=8, color='666666')
    c3.fill = PatternFill('solid', fgColor=GREY)
    c3.alignment = Alignment(horizontal='center')

for col in range(1, 13):
    col_width(ws, col, 13)

# Key findings block
ws.row_dimensions[8].height = 8
section_title(ws, 9, 1, '  KEY FINDINGS & EXAM INTELLIGENCE', span=12)
findings = [
    ('EXAM DIRECTION', 'Injuries/MSK (+1.8/yr) & Reproductive: Female (+1.7/yr) are the fastest-growing body systems. Respiratory (−2.9/yr) and classic Musculoskeletal (−1.8/yr) are declining fastest. The 2025 exam shows a clear pivot toward psychiatric, functional, and reproductive content.'),
    ('DOMINANT QUESTION TYPE', 'Management questions surged +3.3 per year — now the fastest-rising subcategory. Diagnosis (−1.8/yr) and Interpretation (−0.8/yr) are declining. The exam is shifting from "what is it?" toward "what do you do about it?"'),
    ('HIGHEST-YIELD POCKET', 'Cardiovascular × Pharmacology is the single biggest content cell: 59 questions (4.9% of the entire exam). The top 5 combos alone account for 18.3% of all questions.'),
    ('BLUEPRINT TREND', 'Acute Care and Diagnosis has risen from 33.5% (2020) to 37-40% by 2022-2023. Emergent & Urgent Care jumped to 20% in 2024-2025. Preventive Care has declined from 16.5% to 15%.'),
    ('REFERENCE INTELLIGENCE', 'AFP (American Family Physician) dominates: 635 of 2,069 pairs — the single most-cited source. NEJM and Major Journals punch above their weight on Must-Read tier despite lower volume. 2025 references are 100% Unmatched (tier expansion needed).'),
    ('COVERAGE GAPS', 'Sessions 31 (Musculoskeletal Medicine), 35 (ENT), 26 (Neurology), and 30 (Women\'s Health II) have the lowest proportion of tiered references — residents studying these topics should supplement with AFP directly.'),
]
for i, (label, text) in enumerate(findings):
    r = 10 + i
    ws.row_dimensions[r].height = 38
    c = ws.cell(row=r, column=1, value=label)
    c.font = Font(name='Arial', bold=True, size=9, color=WHITE)
    c.fill = PatternFill('solid', fgColor=BLUE)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    c2 = ws.cell(row=r, column=3, value=text)
    c2.font = Font(name='Arial', size=9, color='000000')
    c2.fill = PatternFill('solid', fgColor=GREY if i%2==0 else WHITE)
    c2.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=12)

thin_border(ws, 9, 1, 15, 12)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2: YEAR-OVER-YEAR TRENDS
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('YoY Trends')
ws2.sheet_view.showGridLines = False

ws2.merge_cells('A1:P1')
c = ws2['A1']
c.value = 'YEAR-OVER-YEAR TRENDS  |  Body System, Subcategory & Blueprint'
c.font = Font(name='Arial', bold=True, size=14, color=WHITE)
c.fill = PatternFill('solid', fgColor=NAVY)
c.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 26

# ── Body System × Year raw count ──────────────────────────────────────────────
section_title(ws2, 3, 1, '  BODY SYSTEM — Question Count by Year', span=10)
bs_year = df.groupby(['ExamYear','BodySystem']).size().unstack(fill_value=0)
bs_trend_vals = {}
for col in bs_year.columns:
    slope = np.polyfit(np.arange(len(bs_year)), bs_year[col].values, 1)[0]
    bs_trend_vals[col] = round(float(slope), 2)

yr_cols = list(bs_year.columns)
hdr(ws2, 4, 1, 'Body System', bg=BLUE, halign='left')
for j, yr in enumerate(years):
    hdr(ws2, 4, j+2, str(yr), bg=BLUE)
hdr(ws2, 4, len(years)+2, 'TOTAL', bg=NAVY)
hdr(ws2, 4, len(years)+3, 'Trend/yr', bg=NAVY)
hdr(ws2, 4, len(years)+4, 'Direction', bg=NAVY)

bs_sorted = sorted(bs_trend_vals.items(), key=lambda x: -x[1])
for i, (bs, slope) in enumerate(bs_sorted):
    r = 5 + i
    bg = WHITE if i % 2 == 0 else GREY
    val(ws2, r, 1, bs, bg=bg, halign='left')
    row_total = 0
    for j, yr in enumerate(years):
        n = int(bs_year.loc[yr, bs]) if yr in bs_year.index and bs in bs_year.columns else 0
        val(ws2, r, j+2, n, bg=bg)
        row_total += n
    val(ws2, r, len(years)+2, row_total, bg=bg, bold=True)
    val(ws2, r, len(years)+3, f'{slope:+.2f}', bg=AMBER if slope > 0 else RED_L, bold=True)
    direction = '▲ Rising' if slope > 0.5 else ('▼ Declining' if slope < -0.5 else '→ Stable')
    dir_bg = GREEN if slope > 0.5 else (RED_L if slope < -0.5 else AMBER)
    val(ws2, r, len(years)+4, direction, bg=dir_bg)

thin_border(ws2, 4, 1, 4+len(bs_sorted), len(years)+4)
col_width(ws2, 1, 28)
for c in range(2, len(years)+5):
    col_width(ws2, c, 11)

# ── Subcategory × Year ────────────────────────────────────────────────────────
sc_start = 6 + len(bs_sorted)
section_title(ws2, sc_start, 1, '  SUBCATEGORY — Question Count by Year  (What type of thinking the exam tests)', span=10)
sc_year = df.groupby(['ExamYear','Subcategory']).size().unstack(fill_value=0)
sc_trend_vals = {}
for col in sc_year.columns:
    slope = np.polyfit(np.arange(len(sc_year)), sc_year[col].values, 1)[0]
    sc_trend_vals[col] = round(float(slope), 2)

hdr(ws2, sc_start+1, 1, 'Subcategory', bg=BLUE, halign='left')
for j, yr in enumerate(years):
    hdr(ws2, sc_start+1, j+2, str(yr), bg=BLUE)
hdr(ws2, sc_start+1, len(years)+2, 'TOTAL', bg=NAVY)
hdr(ws2, sc_start+1, len(years)+3, 'Trend/yr', bg=NAVY)
hdr(ws2, sc_start+1, len(years)+4, 'Direction', bg=NAVY)

sc_sorted = sorted(sc_trend_vals.items(), key=lambda x: -x[1])
for i, (sc, slope) in enumerate(sc_sorted):
    r = sc_start + 2 + i
    bg = WHITE if i % 2 == 0 else GREY
    val(ws2, r, 1, sc, bg=bg, halign='left')
    row_total = 0
    for j, yr in enumerate(years):
        n = int(sc_year.loc[yr, sc]) if yr in sc_year.index and sc in sc_year.columns else 0
        val(ws2, r, j+2, n, bg=bg)
        row_total += n
    val(ws2, r, len(years)+2, row_total, bg=bg, bold=True)
    val(ws2, r, len(years)+3, f'{slope:+.2f}', bg=AMBER if slope > 0 else RED_L, bold=True)
    direction = '▲ Rising' if slope > 0.5 else ('▼ Declining' if slope < -0.5 else '→ Stable')
    dir_bg = GREEN if slope > 0.5 else (RED_L if slope < -0.5 else AMBER)
    val(ws2, r, len(years)+4, direction, bg=dir_bg)

thin_border(ws2, sc_start+1, 1, sc_start+1+len(sc_sorted), len(years)+4)

# ── Blueprint × Year ─────────────────────────────────────────────────────────
bp_start = sc_start + len(sc_sorted) + 4
section_title(ws2, bp_start, 1, '  BLUEPRINT CATEGORY — % Share by Year  (ML-predicted, all 1,200 questions)', span=10)
bp_year = df.groupby(['ExamYear','BlueprintCategory_Predicted']).size().unstack(fill_value=0)
bp_pct  = bp_year.div(bp_year.sum(axis=1), axis=0).mul(100).round(1)

hdr(ws2, bp_start+1, 1, 'Blueprint Category', bg=BLUE, halign='left')
for j, yr in enumerate(years):
    hdr(ws2, bp_start+1, j+2, str(yr), bg=BLUE)
hdr(ws2, bp_start+1, len(years)+2, 'Δ 2020→2025', bg=NAVY)

for i, bp in enumerate(bp_pct.columns):
    r = bp_start + 2 + i
    bg = WHITE if i % 2 == 0 else GREY
    val(ws2, r, 1, bp, bg=bg, halign='left')
    for j, yr in enumerate(years):
        pct_val = bp_pct.loc[yr, bp] if yr in bp_pct.index else 0
        c = val(ws2, r, j+2, pct_val, fmt='0.0"%"', bg=bg)
    delta = bp_pct.iloc[-1][bp] - bp_pct.iloc[0][bp]
    val(ws2, r, len(years)+2, f'{delta:+.1f}%', bg=AMBER if delta > 0 else RED_L if delta < 0 else GREY, bold=True)

thin_border(ws2, bp_start+1, 1, bp_start+1+len(bp_pct.columns), len(years)+2)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3: HIGH-YIELD COMBOS
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('HY Combos')
ws3.sheet_view.showGridLines = False

ws3.merge_cells('A1:J1')
c = ws3['A1']
c.value = 'HIGH-YIELD COMBINATIONS  |  BodySystem × Subcategory  |  Top Content Pockets'
c.font = Font(name='Arial', bold=True, size=14, color=WHITE)
c.fill = PatternFill('solid', fgColor=NAVY)
c.alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 26

# Top 20 combos table
section_title(ws3, 3, 1, '  TOP 20 HIGH-YIELD COMBINATIONS  (by total question count, 2020–2025)', span=10)
combo = df.groupby(['BodySystem','Subcategory']).size().reset_index(name='Total')
combo['Pct'] = (combo['Total'] / 1200 * 100).round(2)
combo = combo.sort_values('Total', ascending=False).head(20).reset_index(drop=True)
combo.index += 1

# Add per-year columns
for yr in years:
    yr_df = df[df.ExamYear==yr].groupby(['BodySystem','Subcategory']).size().reset_index(name=str(yr))
    combo = combo.merge(yr_df, on=['BodySystem','Subcategory'], how='left')
    combo[str(yr)] = combo[str(yr)].fillna(0).astype(int)

headers = ['Rank','Body System','Subcategory','Total','% Exam'] + [str(y) for y in years]
bg_list = [NAVY,BLUE,BLUE,NAVY,NAVY] + [BLUE]*6
for j, (h, bg) in enumerate(zip(headers, bg_list)):
    hdr(ws3, 4, j+1, h, bg=bg)

for i, row in combo.iterrows():
    r = 4 + i
    bg = AMBER if i == 1 else (LBLUE if i <= 5 else (WHITE if i % 2 == 0 else GREY))
    val(ws3, r, 1, i, bg=bg, bold=(i<=5))
    val(ws3, r, 2, row['BodySystem'], bg=bg, halign='left', bold=(i<=5))
    val(ws3, r, 3, row['Subcategory'], bg=bg, halign='left')
    val(ws3, r, 4, row['Total'], bg=bg, bold=(i<=5))
    val(ws3, r, 5, row['Pct']/100, fmt='0.00%', bg=bg)
    for j, yr in enumerate(years):
        val(ws3, r, 6+j, int(row[str(yr)]), bg=bg)

thin_border(ws3, 4, 1, 4+20, 5+len(years))
col_width(ws3, 1, 7)
col_width(ws3, 2, 30)
col_width(ws3, 3, 18)
col_width(ws3, 4, 10)
col_width(ws3, 5, 10)
for j in range(len(years)):
    col_width(ws3, 6+j, 9)

# Delta table: biggest movers 2025 vs 2020
delta_start = 28
section_title(ws3, delta_start, 1, '  2025 vs 2020 — BIGGEST MOVERS  (which content pockets gained/lost most)', span=10)

c25 = df[df.ExamYear==2025].groupby(['BodySystem','Subcategory']).size().reset_index(name='2025')
c20 = df[df.ExamYear==2020].groupby(['BodySystem','Subcategory']).size().reset_index(name='2020')
delta = pd.merge(c25, c20, on=['BodySystem','Subcategory'], how='outer').fillna(0)
delta['Delta'] = (delta['2025'] - delta['2020']).astype(int)
delta = delta.sort_values('Delta', ascending=False)

risers  = delta.head(12).reset_index(drop=True)
fallers = delta.tail(12).sort_values('Delta').reset_index(drop=True)

ws3.merge_cells(start_row=delta_start+1, start_column=1, end_row=delta_start+1, end_column=5)
hdr(ws3, delta_start+1, 1, '▲ RISING (2025 vs 2020)', bg=GREEN, fg=NAVY)
hdr(ws3, delta_start+2, 1, 'Body System', bg=BLUE, halign='left')
hdr(ws3, delta_start+2, 2, 'Subcategory', bg=BLUE)
hdr(ws3, delta_start+2, 3, '2020', bg=BLUE)
hdr(ws3, delta_start+2, 4, '2025', bg=BLUE)
hdr(ws3, delta_start+2, 5, 'Δ', bg=NAVY)

for i, row in risers.iterrows():
    r = delta_start + 3 + i
    bg = WHITE if i % 2 == 0 else GREY
    val(ws3, r, 1, row['BodySystem'], bg=bg, halign='left')
    val(ws3, r, 2, row['Subcategory'], bg=bg)
    val(ws3, r, 3, int(row['2020']), bg=bg)
    val(ws3, r, 4, int(row['2025']), bg=bg)
    val(ws3, r, 5, f'+{int(row["Delta"])}', bg=GREEN, bold=True)

hdr(ws3, delta_start+1, 7, '▼ DECLINING (2025 vs 2020)', bg=RED_L, fg=NAVY)
ws3.merge_cells(start_row=delta_start+1, start_column=7, end_row=delta_start+1, end_column=11)
hdr(ws3, delta_start+2, 7, 'Body System', bg=BLUE, halign='left')
hdr(ws3, delta_start+2, 8, 'Subcategory', bg=BLUE)
hdr(ws3, delta_start+2, 9, '2020', bg=BLUE)
hdr(ws3, delta_start+2, 10, '2025', bg=BLUE)
hdr(ws3, delta_start+2, 11, 'Δ', bg=NAVY)

for i, row in fallers.iterrows():
    r = delta_start + 3 + i
    bg = WHITE if i % 2 == 0 else GREY
    val(ws3, r, 7, row['BodySystem'], bg=bg, halign='left')
    val(ws3, r, 8, row['Subcategory'], bg=bg)
    val(ws3, r, 9, int(row['2020']), bg=bg)
    val(ws3, r, 10, int(row['2025']), bg=bg)
    val(ws3, r, 11, str(int(row['Delta'])), bg=RED_L, bold=True)

thin_border(ws3, delta_start+1, 1, delta_start+2+12, 5)
thin_border(ws3, delta_start+1, 7, delta_start+2+12, 11)
col_width(ws3, 6, 2)
for j in [7,8]:
    col_width(ws3, j, 28 if j==7 else 18)
for j in [9,10,11]:
    col_width(ws3, j, 9)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4: REFERENCE TIER ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet('Reference Analysis')
ws4.sheet_view.showGridLines = False

ws4.merge_cells('A1:L1')
c = ws4['A1']
c.value = 'REFERENCE TIER ANALYSIS  |  2,069 Question→Reference Pairs'
c.font = Font(name='Arial', bold=True, size=14, color=WHITE)
c.fill = PatternFill('solid', fgColor=NAVY)
c.alignment = Alignment(horizontal='center', vertical='center')
ws4.row_dimensions[1].height = 26

# Tier by year
section_title(ws4, 3, 1, '  REFERENCE TIER COVERAGE BY YEAR', span=8)
tier_year = qrp.groupby(['ExamYear','Tier']).size().unstack(fill_value=0)
tier_pct  = tier_year.div(tier_year.sum(axis=1), axis=0).mul(100).round(1)
tier_order = ['Must-Read','Core','Supplementary','Unmatched']
tier_order = [t for t in tier_order if t in tier_year.columns]

hdr(ws4, 4, 1, 'Year', bg=BLUE)
hdr(ws4, 4, 2, 'Total Refs', bg=BLUE)
for j, t in enumerate(tier_order):
    hdr(ws4, 4, 3+j, t, bg=NAVY if t=='Must-Read' else BLUE)
    hdr(ws4, 4, 3+len(tier_order)+j, f'{t} %', bg=NAVY if t=='Must-Read' else BLUE)
hdr(ws4, 4, 3+len(tier_order)*2, 'Tier Coverage', bg=NAVY)

for i, yr in enumerate(years):
    r = 5 + i
    bg = AMBER if yr >= 2024 else (WHITE if i%2==0 else GREY)
    total = int(tier_year.loc[yr].sum()) if yr in tier_year.index else 0
    val(ws4, r, 1, yr, bg=bg, bold=True)
    val(ws4, r, 2, total, bg=bg, bold=True)
    tiered = 0
    for j, t in enumerate(tier_order):
        n = int(tier_year.loc[yr, t]) if yr in tier_year.index and t in tier_year.columns else 0
        val(ws4, r, 3+j, n, bg=bg)
        if t != 'Unmatched':
            tiered += n
    for j, t in enumerate(tier_order):
        pct = float(tier_pct.loc[yr, t]) if yr in tier_pct.index and t in tier_pct.columns else 0
        val(ws4, r, 3+len(tier_order)+j, pct/100, fmt='0.0%', bg=bg)
    coverage = tiered/total*100 if total > 0 else 0
    note = '⚠ Tier expansion needed' if coverage < 20 else ('★ Full coverage' if coverage > 80 else '')
    val(ws4, r, 3+len(tier_order)*2, f'{coverage:.0f}% tiered  {note}',
        bg=RED_L if coverage < 20 else (GREEN if coverage > 80 else AMBER))

thin_border(ws4, 4, 1, 4+len(years), 3+len(tier_order)*2)
col_width(ws4, 1, 8); col_width(ws4, 2, 12)
for j in range(len(tier_order)*2+1):
    col_width(ws4, 3+j, 14)

# Source type breakdown
src_start = 5 + len(years) + 3
section_title(ws4, src_start, 1, '  REFERENCE SOURCE TYPE ANALYSIS  (AFP dominates; NEJM/Journals punch above weight on Must-Read)', span=9)

def classify_ref(ref):
    r = str(ref).lower()
    if 'uspstf' in r or 'u.s. preventive' in r: return 'USPSTF'
    if 'cdc' in r or 'centers for disease' in r: return 'CDC'
    if 'aafp' in r or 'american academy of fam' in r: return 'AAFP'
    if 'aha' in r or 'acc' in r or 'american heart' in r or 'american college of card' in r: return 'AHA/ACC'
    if 'ada ' in r or 'american diabetes' in r: return 'ADA'
    if 'acog' in r or 'american college of ob' in r: return 'ACOG'
    if 'idsa' in r or 'infectious disease' in r: return 'IDSA'
    if 'am fam physician' in r or ' afp ' in r: return 'AFP'
    if 'n engl j med' in r or 'nejm' in r: return 'NEJM'
    if any(x in r for x in ['jama','lancet','bmj','ann intern med','annals']): return 'Major Journal'
    if any(x in r for x in ['harrison','tintinalli','lange','uptodate']): return 'Textbook'
    return 'Specialty Journal'

qrp['source_type'] = qrp['RefMatched'].apply(classify_ref)
src = qrp.groupby(['source_type','Tier']).size().unstack(fill_value=0).reset_index()
src['total'] = src[[c for c in src.columns if c != 'source_type']].sum(axis=1)
src = src.sort_values('total', ascending=False)
tier_cols_src = [t for t in tier_order if t in src.columns]

hdr(ws4, src_start+1, 1, 'Source Type', bg=BLUE, halign='left')
hdr(ws4, src_start+1, 2, 'Total', bg=NAVY)
for j, t in enumerate(tier_cols_src):
    hdr(ws4, src_start+1, 3+j, t, bg=BLUE)
hdr(ws4, src_start+1, 3+len(tier_cols_src), '% Must-Read+Core', bg=NAVY)

for i, row in src.iterrows():
    r = src_start + 2 + list(src.index).index(i)
    bg = AMBER if row['source_type'] == 'AFP' else (WHITE if r%2==0 else GREY)
    val(ws4, r, 1, row['source_type'], bg=bg, halign='left', bold=(row['source_type']=='AFP'))
    val(ws4, r, 2, int(row['total']), bg=bg, bold=True)
    tiered_n = 0
    for j, t in enumerate(tier_cols_src):
        n = int(row[t]) if t in row.index else 0
        val(ws4, r, 3+j, n, bg=bg)
        if t in ('Must-Read','Core'): tiered_n += n
    pct_t = tiered_n / int(row['total']) if row['total'] > 0 else 0
    val(ws4, r, 3+len(tier_cols_src), pct_t, fmt='0%', bg=GREEN if pct_t > 0.7 else bg)

thin_border(ws4, src_start+1, 1, src_start+1+len(src), 3+len(tier_cols_src))
col_width(ws4, 1, 20)
for j in range(len(tier_cols_src)+2):
    col_width(ws4, 2+j, 14)

# Top cited references
top_start = src_start + len(src) + 4
section_title(ws4, top_start, 1, '  TOP 25 MOST-CITED REFERENCES  (across all years)', span=9)
ref_counts = qrp[qrp['RefMatched'].str.len() > 20].groupby('RefMatched').agg(
    Citations=('QuestionID','count'),
    Tier=('Tier','first'),
    Years=('ExamYear', lambda x: len(x.unique()))
).reset_index().sort_values('Citations', ascending=False).head(25).reset_index(drop=True)
ref_counts.index += 1

hdr(ws4, top_start+1, 1, '#', bg=BLUE)
hdr(ws4, top_start+1, 2, 'Reference', bg=BLUE, halign='left')
hdr(ws4, top_start+1, 3, 'Citations', bg=NAVY)
hdr(ws4, top_start+1, 4, 'Tier', bg=NAVY)
hdr(ws4, top_start+1, 5, 'Yrs Seen', bg=BLUE)

for i, row in ref_counts.iterrows():
    r = top_start + 1 + i
    tier_bg = AMBER if row['Tier']=='Must-Read' else (LBLUE if row['Tier']=='Core' else (WHITE if i%2==0 else GREY))
    val(ws4, r, 1, i, bg=tier_bg)
    val(ws4, r, 2, row['RefMatched'][:120], bg=tier_bg, halign='left')
    val(ws4, r, 3, int(row['Citations']), bg=tier_bg, bold=True)
    val(ws4, r, 4, row['Tier'], bg=tier_bg)
    val(ws4, r, 5, int(row['Years']), bg=tier_bg)
    ws4.row_dimensions[r].height = 28

thin_border(ws4, top_start+1, 1, top_start+1+25, 5)
col_width(ws4, 1, 6); col_width(ws4, 2, 70); col_width(ws4, 3, 12)
col_width(ws4, 4, 15); col_width(ws4, 5, 10)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5: SESSION COVERAGE
# ══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet('Session Coverage')
ws5.sheet_view.showGridLines = False

ws5.merge_cells('A1:M1')
c = ws5['A1']
c.value = 'AAFP SESSION COVERAGE QUALITY  |  48 Sessions  |  238 Questions Placed  |  387 References'
c.font = Font(name='Arial', bold=True, size=14, color=WHITE)
c.fill = PatternFill('solid', fgColor=NAVY)
c.alignment = Alignment(horizontal='center', vertical='center')
ws5.row_dimensions[1].height = 26

section_title(ws5, 3, 1, '  SESSION-LEVEL COVERAGE DETAIL  (sorted by session number)', span=13)

sess_rows = []
for sid, s in v6.items():
    refs  = s['refs']
    tiers = [r['tier'] for r in refs]
    scores = [q['kw_score'] for q in s['questions']]
    yr_ct = {}
    for q in s['questions']:
        yr_ct[q['year']] = yr_ct.get(q['year'],0)+1
    must_r = tiers.count('Must-Read')
    core_n = tiers.count('Core')
    unm    = tiers.count('Unmatched')
    tot    = len(refs)
    pct_t  = (must_r+core_n)/max(tot,1)*100
    sess_rows.append({
        'ID': int(sid),
        'Session': s['session_title'],
        'Qs': s['question_count'],
        'Refs': tot,
        'Must-Read': must_r,
        'Core': core_n,
        'Unmatched': unm,
        '% Tiered': pct_t,
        'Avg Score': round(np.mean(scores) if scores else 0, 2),
        **{str(yr): yr_ct.get(yr,0) for yr in years},
    })
sess_df = pd.DataFrame(sess_rows).sort_values('ID').reset_index(drop=True)

h_list = ['#','Session Title','Qs','Refs','★ Must-Read','Core','Unmatched','% Tiered','Avg Match'] + [str(y) for y in years]
bg_list = [NAVY,BLUE,BLUE,BLUE,NAVY,BLUE,BLUE,NAVY,BLUE]+[BLUE]*6
for j,(h,bg) in enumerate(zip(h_list,bg_list)):
    hdr(ws5, 4, j+1, h, bg=bg, sz=9)

for i, row in sess_df.iterrows():
    r = 5 + i
    pct = row['% Tiered']
    row_bg = RED_L if pct < 20 else (AMBER if pct < 50 else (GREEN if pct >= 75 else WHITE if i%2==0 else GREY))
    val(ws5, r, 1, int(row['ID']), bg=row_bg, bold=True)
    val(ws5, r, 2, row['Session'], bg=row_bg, halign='left')
    val(ws5, r, 3, int(row['Qs']), bg=row_bg)
    val(ws5, r, 4, int(row['Refs']), bg=row_bg)
    val(ws5, r, 5, int(row['Must-Read']), bg=AMBER if row['Must-Read']>0 else row_bg, bold=row['Must-Read']>0)
    val(ws5, r, 6, int(row['Core']), bg=row_bg)
    val(ws5, r, 7, int(row['Unmatched']), bg=RED_L if row['Unmatched']>3 else row_bg)
    val(ws5, r, 8, pct/100, fmt='0%', bg=row_bg, bold=True)
    val(ws5, r, 9, row['Avg Score'], bg=row_bg)
    for j, yr in enumerate(years):
        n = int(row[str(yr)])
        val(ws5, r, 10+j, n, bg=row_bg)

thin_border(ws5, 4, 1, 4+len(sess_df), len(h_list))
col_width(ws5, 1, 5); col_width(ws5, 2, 42); col_width(ws5, 3, 6)
col_width(ws5, 4, 7); col_width(ws5, 5, 12); col_width(ws5, 6, 8)
col_width(ws5, 7, 11); col_width(ws5, 8, 10); col_width(ws5, 9, 12)
for j in range(6):
    col_width(ws5, 10+j, 8)

# Legend
leg_r = 5 + len(sess_df) + 2
ws5.merge_cells(start_row=leg_r, start_column=1, end_row=leg_r, end_column=4)
c = ws5.cell(row=leg_r, column=1, value='COLOR LEGEND:')
c.font = Font(name='Arial', bold=True, size=9, color=NAVY)

legend = [('RED', RED_L,'< 20% tiered — supplement with AFP directly'),
          ('AMBER', AMBER,'20–49% tiered — good base, some gaps'),
          ('WHITE/GREY', GREY,'50–74% tiered — solid coverage'),
          ('GREEN', GREEN,'≥ 75% tiered — excellent reference quality')]
for k,(label,bg,desc) in enumerate(legend):
    ws5.cell(row=leg_r+1+k, column=1, value=label).fill = PatternFill('solid', fgColor=bg)
    ws5.cell(row=leg_r+1+k, column=1).font = Font(name='Arial', size=9, bold=True)
    ws5.merge_cells(start_row=leg_r+1+k, start_column=2, end_row=leg_r+1+k, end_column=6)
    ws5.cell(row=leg_r+1+k, column=2, value=desc).font = Font(name='Arial', size=9)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 6: RESIDENT STUDY GUIDE
# ══════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet('Resident Study Guide')
ws6.sheet_view.showGridLines = False

ws6.merge_cells('A1:J1')
c = ws6['A1']
c.value = 'RESIDENT STUDY GUIDE  |  Prioritized by ITE Exam Evidence  |  2020–2025 Analysis'
c.font = Font(name='Arial', bold=True, size=14, color=WHITE)
c.fill = PatternFill('solid', fgColor=NAVY)
c.alignment = Alignment(horizontal='center', vertical='center')
ws6.row_dimensions[1].height = 26

# Priority tiers
section_title(ws6, 3, 1, '  STUDY PRIORITY FRAMEWORK', span=10)
priority_data = [
    ('TIER 1: MASTER FIRST','Pharmacology in high-volume body systems. These combos alone cover ~18% of the exam.',
     'Cardiovascular × Pharm (59q), Respiratory × Pharm (52q), Endocrine × Pharm (40q), Psychogenic × Pharm (29q), MSK × Pharm (27q)'),
    ('TIER 2: HIGH-YIELD MANAGEMENT','Management questions are rising fastest (+3.3/yr). Board now tests "what you do" more than "what it is".',
     'Cardiovascular × Mgmt (38q), Respiratory × Mgmt (28q), MSK × Mgmt (23q), Nonspecific × Mgmt (21q)'),
    ('TIER 3: GROWING AREAS','Content pockets growing fastest — expect heavier weight in future exams.',
     'Injuries/MSK (+1.8/yr), Reproductive: Female (+1.7/yr), Psychiatric/Behavioral (+1.2/yr), Neurologic (+1.1/yr)'),
    ('TIER 4: SCREENING & PREVENTION','Population-Based Care × Screening is #5 overall (31q). USPSTF recs are consistently cited.',
     'Pop-Based Care × Screening (31q), Respiratory × Prevention (16q). Read USPSTF A/B recommendations.'),
    ('TIER 5: WATCH — DECLINING','These areas are shrinking on the exam but still sizable. Maintain but do not over-invest.',
     'Classic Respiratory (−2.9/yr), Musculoskeletal traditional (−1.8/yr), Endocrine (−1.5/yr)'),
]
for i,(tier,rationale,focus) in enumerate(priority_data):
    r_base = 4 + i*4
    bg = [AMBER,LBLUE,GREEN,GREY,RED_L][i]
    ws6.merge_cells(start_row=r_base, start_column=1, end_row=r_base, end_column=10)
    c = ws6.cell(row=r_base, column=1, value=tier)
    c.font = Font(name='Arial', bold=True, size=11, color=NAVY)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws6.row_dimensions[r_base].height = 18

    ws6.merge_cells(start_row=r_base+1, start_column=1, end_row=r_base+1, end_column=10)
    c2 = ws6.cell(row=r_base+1, column=1, value=rationale)
    c2.font = Font(name='Arial', size=9, italic=True, color='444444')
    c2.fill = PatternFill('solid', fgColor=WHITE)
    c2.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws6.row_dimensions[r_base+1].height = 28

    ws6.merge_cells(start_row=r_base+2, start_column=1, end_row=r_base+2, end_column=10)
    c3 = ws6.cell(row=r_base+2, column=1, value='→  ' + focus)
    c3.font = Font(name='Arial', size=10, bold=True, color=NAVY)
    c3.fill = PatternFill('solid', fgColor=GREY)
    c3.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws6.row_dimensions[r_base+2].height = 28

for j in range(1, 11):
    col_width(ws6, j, 16 if j==1 else 12)
col_width(ws6, 1, 30)

# Reference reading list
rr_start = 26
section_title(ws6, rr_start, 1, '  PRIORITY READING LIST  (AFP first — single highest-yield source)', span=10)
reading = [
    ('AFP', 'American Family Physician', '635 citations in 6-year QRP — #1 source by volume. Read AFP editorials on Pharmacology & Management topics.'),
    ('NEJM', 'New England Journal of Medicine', '67 citations, all Core/Must-Read tier. Highest tier-density of any source. Read landmark trials cited in Cardiology/Respiratory.'),
    ('AHA/ACC', 'Cardiology Guidelines', '62 citations — Cardiovascular × Pharmacology is the #1 exam pocket. Know 2023 ACC/AHA HTN, HF, lipid guidelines.'),
    ('CDC', 'Centers for Disease Control', '49 citations, 96% Core. Vaccination schedules, STI treatment, infectious disease protocols.'),
    ('USPSTF', 'US Preventive Services Task Force', '5 citations but 100% Core tier. Know all A and B screening recommendations — high-yield per page read.'),
    ('Major Journals','JAMA, Lancet, BMJ, Annals', '88 citations, 82% Core/Must-Read. High density per citation. Focus on meta-analyses cited in top combo pockets.'),
    ('IDSA', 'Infectious Disease Society', '16 citations, 94% Core. ID management guidelines especially for STIs, skin/soft tissue, respiratory infections.'),
]
hdr(ws6, rr_start+1, 1, 'Source', bg=NAVY)
hdr(ws6, rr_start+1, 2, 'Full Name', bg=BLUE)
hdr(ws6, rr_start+1, 3, 'Study Rationale', bg=BLUE, halign='left')
ws6.merge_cells(start_row=rr_start+1, start_column=3, end_row=rr_start+1, end_column=10)

for i,(abbr, name, rationale) in enumerate(reading):
    r = rr_start + 2 + i
    bg = AMBER if abbr=='AFP' else (WHITE if i%2==0 else GREY)
    ws6.row_dimensions[r].height = 30
    val(ws6, r, 1, abbr, bg=bg, bold=True)
    val(ws6, r, 2, name, bg=bg, halign='left')
    ws6.merge_cells(start_row=r, start_column=3, end_row=r, end_column=10)
    c = ws6.cell(row=r, column=3, value=rationale)
    c.font = Font(name='Arial', size=9)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

thin_border(ws6, rr_start+1, 1, rr_start+1+len(reading), 10)
col_width(ws6, 1, 14); col_width(ws6, 2, 30)

wb.save(str(OUT))
print(f'Saved: {OUT}')
print(f'Sheets: {[s.title for s in wb.worksheets]}')
