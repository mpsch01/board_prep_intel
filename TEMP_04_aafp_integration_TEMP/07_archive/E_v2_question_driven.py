"""
E_v2_question_driven.py
=======================
Question-driven session matching with inline ref extraction.

Workflow:
  1. Load keyword library (48 sessions, scored terms)
  2. Load AI_Tagged xlsx (1200 questions, refs in Explanation column)
  3. For each question:
       a. Tokenize QuestionStem + AI_ClinicalFocus + Explanation
       b. Score keyword overlap against every session → assign to best session(s)
  4. Per session: collect top-N questions by score
  5. For each assigned question: parse refs from Explanation → fuzzy-match tier DB
  6. Per session: deduplicate refs, label tier, rank
  7. Output: session_hy_inserts_v5.json

Tuning:
  MAX_Q_PER_SESSION = 5      top questions shown per session
  MIN_KW_SCORE      = 0.10   min composite overlap to consider
  MIN_KW_HITS       = 2      min distinct keyword matches required
  MATCH_THRESHOLD   = 0.70   fuzzy match cutoff for ref → tier DB
"""

import json, csv, re, os
from openpyxl import load_workbook
from difflib import SequenceMatcher
from collections import defaultdict

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE     = r'C:\Users\mpsch\Desktop\claude_knowledge\board_prep'
KW_JSON  = os.path.join(BASE, 'keyword_library', 'session_keyword_library.json')
XLSX     = os.path.join(BASE, 'ite_exam', '03_database', 'ABFM_ITE_AI_Tagged_excel.xlsx')
TIER_CSV = os.path.join(BASE, 'ite_refs', '02_working', 'ITE_Reference_Tiers_Clean.csv')
CW_CSV   = os.path.join(BASE, 'aafp_integration', '02_working', 'session_cluster_crosswalk.csv')
OUT_JSON = os.path.join(BASE, 'aafp_integration', '02_working', 'session_hy_inserts_v5.json')
VAL_CSV  = os.path.join(BASE, 'keyword_library', 'raw_files', 'E_v2_validation.csv')

# ── Tuning ─────────────────────────────────────────────────────────────────────
MAX_Q_PER_SESSION = 5
MIN_KW_SCORE      = 0.10
MIN_KW_HITS       = 2
MATCH_THRESHOLD   = 0.70

# ── Stopwords ──────────────────────────────────────────────────────────────────
STOPWORDS = {
    'the','a','an','of','in','to','and','or','is','are','was','were','be',
    'been','being','have','has','had','do','does','did','will','would','could',
    'should','may','might','can','this','that','these','those','with','for',
    'from','not','but','at','by','as','it','its','he','she','they','their',
    'which','who','what','when','where','how','one','two','three','also',
    'most','more','other','following','patient','patients','which','year',
    'old','male','female','history','presents','exam','test','use','used',
    'using','associated','include','including','such','based','than','both',
    'after','before','during','without','within','per','type','level','rate',
    'new','well','age','used','first','initial','given','known','seen',
    'risk','blood','pressure','treatment','diagnosis','management','clinical',
}

def tokenize(text):
    """Lowercase, split on non-alpha, remove stopwords, min 3 chars."""
    if not text:
        return set()
    tokens = re.findall(r'[a-zA-Z]{3,}', text.lower())
    return {t for t in tokens if t not in STOPWORDS}

def tokenize_bigrams(text):
    """Return unigrams + bigrams."""
    tokens = re.findall(r'[a-zA-Z]{3,}', (text or '').lower())
    filtered = [t for t in tokens if t not in STOPWORDS]
    unigrams = set(filtered)
    bigrams  = {f'{filtered[i]} {filtered[i+1]}' for i in range(len(filtered)-1)}
    return unigrams | bigrams

# ── Load keyword library ───────────────────────────────────────────────────────
def load_keyword_library(path):
    with open(path, encoding='utf-8') as f:
        raw = json.load(f)
    sessions = {}
    for sess_id, data in raw.items():
        kw_list = data.get('keywords', [])
        # Build dict: term → composite score
        terms = {}
        for kw in kw_list:
            t = kw['term'].lower()
            terms[t] = kw['composite']
        sessions[sess_id] = {
            'session_name': data.get('session_name', ''),
            'terms': terms,
            'max_composite': max(terms.values()) if terms else 1.0
        }
    return sessions

# ── Load tier DB ───────────────────────────────────────────────────────────────
def load_tier_db(path):
    rows = []
    with open(path, newline='', encoding='utf-8') as f:
        for row in csv.DictReader(f):
            norm = re.sub(r'[^\w\s]', ' ', row['CleanRef'].lower())
            norm = re.sub(r'\s+', ' ', norm).strip()
            rows.append({
                'CleanRef': row['CleanRef'],
                'Tier':     row['Tier'],
                'norm':     norm
            })
    return rows

def fuzzy_match_ref(raw, tier_db):
    norm_raw = re.sub(r'[^\w\s]', ' ', raw.lower())
    norm_raw = re.sub(r'\s+', ' ', norm_raw).strip()
    best_score, best = 0.0, None
    for row in tier_db:
        s = SequenceMatcher(None, norm_raw[:200], row['norm'][:200]).ratio()
        if s > best_score:
            best_score, best = s, row
    if best_score >= MATCH_THRESHOLD and best:
        return best['CleanRef'], best['Tier'], round(best_score, 3)
    return raw.strip(), 'Unmatched', round(best_score, 3)

# ── Parse refs from Explanation ────────────────────────────────────────────────
REF_INLINE  = re.compile(r'Ref:\s*(.+)', re.DOTALL | re.IGNORECASE)
REF_HEADING = re.compile(r'References?\s*\n(.+)', re.DOTALL | re.IGNORECASE)
NUM_SPLIT   = re.compile(r'\s+\d+\)\s+')
CITE_HINT   = re.compile(r'\b(19|20)\d{2}\b')

def parse_refs_from_explanation(expl):
    """Extract raw citation strings from Explanation field."""
    if not expl:
        return []
    refs = []

    # 2020-2023 style: "Ref: Author...: Title. Journal Year;..."
    m = REF_INLINE.search(expl)
    if m:
        ref_block = m.group(1).strip()
        # Split on numbered items: " 2) ", " 3) "
        parts = NUM_SPLIT.split(ref_block)
        for p in parts:
            p = p.strip().rstrip('0123456789').strip()
            if len(p) > 20 and CITE_HINT.search(p):
                refs.append(p)
        return refs

    # 2024 style: "Reference\nAuthor. Title. Journal Year;..."
    m2 = REF_HEADING.search(expl)
    if m2:
        block = m2.group(1).strip()
        # Split on blank lines or new author lines
        lines = [l.strip() for l in block.split('\n') if l.strip()]
        current = ''
        for line in lines:
            is_new = bool(re.match(r'^[A-Z][a-zA-Z]', line)) and bool(CITE_HINT.search(line))
            if is_new:
                if current and len(current) > 20:
                    refs.append(current.strip())
                current = line
            else:
                current += ' ' + line
        if current and len(current) > 20:
            refs.append(current.strip())

    return refs

# ── Load questions from xlsx ───────────────────────────────────────────────────
def load_questions(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() if h else '' for h in rows[0]]
    questions = []
    for row in rows[1:]:
        d = dict(zip(headers, row))
        questions.append({
            'qid':        str(d.get('Question ID', '') or ''),
            'year':       int(d.get('ExamYear', 0) or 0),
            'stem':       str(d.get('QuestionStem', '') or ''),
            'explanation':str(d.get('Explanation', '') or ''),
            'ai_focus':   str(d.get('AI_ClinicalFocus', '') or ''),
            'ai_subcat':  str(d.get('AI_Subcategory', '') or ''),
            'category':   str(d.get('PrimaryCategory', '') or ''),
        })
    wb.close()
    return questions

# ── Score one question against one session ─────────────────────────────────────
def score_question_vs_session(q_tokens, session):
    """Returns (composite_score, hit_count)."""
    terms = session['terms']
    max_c = session['max_composite']
    score = 0.0
    hits  = 0
    for tok in q_tokens:
        if tok in terms:
            score += terms[tok] / max_c
            hits  += 1
        # check bigram: tok might be 'tok1 tok2'
    # bigrams already in q_tokens from tokenize_bigrams
    return round(score, 4), hits

# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    print('Loading keyword library...')
    sessions = load_keyword_library(KW_JSON)
    print(f'  {len(sessions)} sessions loaded')

    print('Loading tier database...')
    tier_db = load_tier_db(TIER_CSV)
    print(f'  {len(tier_db)} refs in tier DB')

    print('Loading crosswalk...')
    with open(CW_CSV, newline='', encoding='utf-8') as f:
        cw_rows = list(csv.DictReader(f))
    cw = {r['session_number']: r for r in cw_rows}

    print('Loading questions...')
    questions = load_questions(XLSX)
    print(f'  {len(questions)} questions loaded')

    # ── Score every question against every session ─────────────────────────────
    print('\nScoring questions vs sessions...')
    # session_id → list of (score, hits, question_dict)
    session_matches = defaultdict(list)

    for q in questions:
        # Build token set from stem + AI focus + AI subcat (NOT explanation — too noisy)
        q_text  = f"{q['stem']} {q['ai_focus']} {q['ai_subcat']}"
        q_tokens = tokenize_bigrams(q_text)

        best_score, best_hits, best_sess = 0.0, 0, None
        second_score, second_sess = 0.0, None

        for sess_id, sess_data in sessions.items():
            sc, hits = score_question_vs_session(q_tokens, sess_data)
            if sc > best_score:
                second_score, second_sess = best_score, best_sess
                best_score, best_hits, best_sess = sc, hits, sess_id
            elif sc > second_score:
                second_score, second_sess = sc, sess_id

        # Primary assignment
        if best_sess and best_score >= MIN_KW_SCORE and best_hits >= MIN_KW_HITS:
            session_matches[best_sess].append((best_score, best_hits, q))

        # Secondary assignment if strong enough (within 20% of best)
        if (second_sess and second_score >= MIN_KW_SCORE
                and best_hits >= MIN_KW_HITS
                and second_score >= best_score * 0.80):
            session_matches[second_sess].append((second_score, best_hits, q))

    # ── Per session: top N questions → pull refs ───────────────────────────────
    print('Building session inserts...')
    output   = {}
    val_rows = []  # for validation CSV

    tier_order = {'Must-Read': 0, 'Core': 1, 'Supplementary': 2, 'Unmatched': 3}

    for sess_id in sorted(sessions.keys()):
        matches = session_matches.get(sess_id, [])
        # Sort by score desc, deduplicate by qid
        seen_qids = set()
        ranked = []
        for sc, hits, q in sorted(matches, key=lambda x: -x[0]):
            if q['qid'] not in seen_qids:
                seen_qids.add(q['qid'])
                ranked.append((sc, hits, q))

        top_q = ranked[:MAX_Q_PER_SESSION]

        # Collect refs from all top questions
        ref_pool = {}  # CleanRef → {'tier', 'score', 'qids'}
        for sc, hits, q in top_q:
            raw_refs = parse_refs_from_explanation(q['explanation'])
            for raw in raw_refs:
                matched, tier, mscore = fuzzy_match_ref(raw, tier_db)
                if matched not in ref_pool:
                    ref_pool[matched] = {'tier': tier, 'match_score': mscore, 'qids': []}
                ref_pool[matched]['qids'].append(q['qid'])

        # Sort refs: Must-Read first, then Core, Supplementary, Unmatched
        sorted_refs = sorted(
            ref_pool.items(),
            key=lambda x: (tier_order.get(x[1]['tier'], 9), -x[1]['match_score'])
        )

        sess_name = sessions[sess_id]['session_name']
        cw_info   = cw.get(str(int(sess_id)), {})

        output[sess_id] = {
            'session_id':    sess_id,
            'session_title': sess_name,
            'question_count': len(top_q),
            'questions': [
                {
                    'qid':         q['qid'],
                    'year':        q['year'],
                    'focus':       q['ai_focus'],
                    'stem_preview': q['stem'][:150],
                    'kw_score':    sc,
                    'kw_hits':     hits,
                }
                for sc, hits, q in top_q
            ],
            'refs': [
                {
                    'citation':     ref,
                    'tier':         info['tier'],
                    'match_score':  info['match_score'],
                    'cited_by':     info['qids'],
                }
                for ref, info in sorted_refs
            ],
            'must_read_count':    sum(1 for _, i in sorted_refs if i['tier'] == 'Must-Read'),
            'core_count':         sum(1 for _, i in sorted_refs if i['tier'] == 'Core'),
            'supplementary_count':sum(1 for _, i in sorted_refs if i['tier'] == 'Supplementary'),
        }

        # Validation rows
        for sc, hits, q in top_q:
            val_rows.append({
                'session_id':   sess_id,
                'session_title':sess_name,
                'qid':          q['qid'],
                'year':         q['year'],
                'kw_score':     sc,
                'kw_hits':      hits,
                'ai_focus':     q['ai_focus'],
            })

    # ── Write outputs ──────────────────────────────────────────────────────────
    print(f'\nWriting {OUT_JSON}...')
    with open(OUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(output, f, indent=2)
    print(f'  {len(output)} sessions written')

    print(f'Writing {VAL_CSV}...')
    os.makedirs(os.path.dirname(VAL_CSV), exist_ok=True)
    with open(VAL_CSV, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=['session_id','session_title','qid','year','kw_score','kw_hits','ai_focus'])
        w.writeheader()
        w.writerows(val_rows)

    # ── Summary ────────────────────────────────────────────────────────────────
    print('\n=== SUMMARY ===')
    total_q    = sum(v['question_count'] for v in output.values())
    total_refs = sum(len(v['refs']) for v in output.values())
    total_mr   = sum(v['must_read_count'] for v in output.values())
    total_core = sum(v['core_count'] for v in output.values())
    no_q       = sum(1 for v in output.values() if v['question_count'] == 0)
    no_mr      = sum(1 for v in output.values() if v['must_read_count'] == 0)

    print(f'  Sessions with questions : {len(output)-no_q}/48')
    print(f'  Sessions with must-read : {len(output)-no_mr}/48')
    print(f'  Total questions placed  : {total_q}')
    print(f'  Total unique refs placed: {total_refs}')
    print(f'  Must-Read refs placed   : {total_mr}')
    print(f'  Core refs placed        : {total_core}')
    print(f'  Sessions w/ 0 questions : {no_q}')

    print('\nSpot-check Session 02 (PVD):')
    s02 = output.get('02', {})
    print(f'  Questions: {s02.get("question_count",0)}')
    for q in s02.get('questions', []):
        print(f'    [{q["year"]}] {q["focus"]} (score={q["kw_score"]}, hits={q["kw_hits"]})')
    print(f'  Must-Read refs: {s02.get("must_read_count",0)}')
    for r in s02.get('refs', [])[:5]:
        print(f'    [{r["tier"]}] {r["citation"][:100]}')

    print('\nDone.')

if __name__ == '__main__':
    main()
