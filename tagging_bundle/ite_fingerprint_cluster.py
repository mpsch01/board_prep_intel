"""
ITE Exam Fingerprint Clustering Pipeline
=========================================
TF-IDF + K-Means + UMAP on 1,200 ITE question stems.
Surfaces recurring exam archetypes -- HOW ABFM tests, not just WHAT.

Usage:
    python ite_fingerprint_cluster.py

Output:
    ite_exam/04_outputs/ITE_ExamFingerprint_Clustered.xlsx
    ite_exam/04_outputs/ITE_Fingerprint_cluster_summary.json
"""

import re, json, warnings, os
import numpy as np
import pandas as pd
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.cluster import KMeans
from sklearn.metrics import silhouette_score
from sklearn.decomposition import TruncatedSVD
from sklearn.preprocessing import normalize
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

BASE = r'C:\Users\mpsch\Desktop\claude_knowledge\board_prep\ite_exam'
MASTER_PATH = os.path.join(BASE, '03_database', 'ABFM_ITE_Master_v2.xlsx')
OUT_XLSX = os.path.join(BASE, '04_outputs', 'ITE_ExamFingerprint_Clustered.xlsx')
OUT_JSON = os.path.join(BASE, '04_outputs', 'ITE_Fingerprint_cluster_summary.json')

# ── CLINICAL STOPWORDS ──────────────────────────────────────────────────────
CLINICAL_STOPWORDS = {
    'patient','year','old','presents','history','physical','exam',
    'examination','following','best','appropriate','next','step',
    'likely','most','management','diagnosis','treatment','finding',
    'results','shows','reveals','noted','reported','complaint',
    'including','currently','recent','new','known','past','medical',
    'medication','medications','taking','prescribed','received',
    'weeks','months','days','hours','ago','states','denies',
    'office','clinic','hospital','emergency','visit','evaluation',
    'review','systems','vital','signs','normal','unremarkable',
    'laboratory','blood','level','levels','test','testing',
    'negative','positive','elevated','decreased','increased',
    'woman','man','female','male','girl','boy','child','adult',
    'physician','doctor','clinician','provider','recommend',
    'well','also','since','two','three','four','five','six',
    'would','could','should','one','following','which','what',
}

# ── ARCHETYPE RULES ──────────────────────────────────────────────────────────
ARCHETYPE_RULES = [
    ('DIAGNOSIS',      r'\b(?:most likely diagnosis|diagnosis is|confirms diagnosis|consistent with|identify the|most likely cause)\b'),
    ('NEXT_STEP',      r'\b(?:next step|next best|most appropriate next|immediate|emergent|urgent action|what should you do)\b'),
    ('MANAGEMENT',     r'\b(?:best management|treatment of choice|appropriate management|how (?:should|would) you manage|initiate|start therapy|first.?line)\b'),
    ('SCREENING',      r'\b(?:screen(?:ing)?|preventive|prevention|uspstf|recommend screening|when to screen|preventive services)\b'),
    ('COUNSELING',     r'\b(?:counsel(?:ing)?|advise|advice|educate|patient education|inform the patient|what (?:should|would) you tell)\b'),
    ('PHARMACOLOGY',   r'\b(?:medication|drug|prescribe|dose|dosing|side effect|adverse|contraindicated|interaction|antibiotic|statin|anti.?hypertensive|pharmacotherapy)\b'),
    ('INTERPRETATION', r'\b(?:interpret|result shows|finding|laboratory|ecg|imaging|biopsy|culture|demonstrates|reveals|which (?:finding|result))\b'),
    ('REFERRAL',       r'\b(?:refer(?:ral)?|consult|specialist|admit|hospitalize|transfer|escalate)\b'),
    ('SURVEILLANCE',   r'\b(?:follow.?up|monitoring|surveillance|repeat|interval|long.?term|ongoing|re.?check)\b'),
    ('RISK_FACTORS',   r'\b(?:risk factor|risk for|predispose|association|most associated|etiology|cause of|pathophysiology)\b'),
]

ARCHETYPE_DISPLAY = {
    'DIAGNOSIS':      '🔬 Diagnosis',
    'NEXT_STEP':      '⚡ Next Step',
    'MANAGEMENT':     '💊 Management',
    'SCREENING':      '🛡️ Screening',
    'COUNSELING':     '💬 Counseling',
    'PHARMACOLOGY':   '💉 Pharmacology',
    'INTERPRETATION': '📊 Interpretation',
    'REFERRAL':       '🏥 Referral/Admit',
    'SURVEILLANCE':   '🔄 Surveillance',
    'RISK_FACTORS':   '⚠️ Risk Factors',
    'UNCLASSIFIED':   '❓ Mixed',
}

def clean_stem(text):
    if not isinstance(text, str): return ''
    text = text.lower()
    text = re.sub(r'\b[a-e]\.\s+[^\n]+', ' ', text)  # remove answer choices
    text = re.sub(r'\b\d+\b', ' ', text)
    text = re.sub(r'[^\w\s]', ' ', text)
    return re.sub(r'\s+', ' ', text).strip()

def detect_archetype(text):
    if not isinstance(text, str): return 'UNCLASSIFIED'
    t = text.lower()
    for label, pat in ARCHETYPE_RULES:
        if re.search(pat, t): return label
    return 'UNCLASSIFIED'

# ── LOAD DATA ────────────────────────────────────────────────────────────────
print("Loading master question bank...")
master = pd.read_excel(MASTER_PATH)
print(f"  {len(master)} questions loaded")

df = master[['QuestionID','ExamYear','QuestionStem','BodySystem','Subcategory',
             'BlueprintCategory_Predicted','CorrectAnswer']].copy()
df['stem_clean'] = df['QuestionStem'].apply(clean_stem)
df['rule_archetype'] = df['QuestionStem'].apply(detect_archetype)

# ── TF-IDF + LSA ─────────────────────────────────────────────────────────────
print("Building TF-IDF matrix...")
vectorizer = TfidfVectorizer(
    max_features=2500,
    ngram_range=(1, 2),
    sublinear_tf=True,
    min_df=3,
    max_df=0.85,
    stop_words='english',
    token_pattern=r'\b[a-z][a-z]+\b'
)
X = vectorizer.fit_transform(df['stem_clean'])
print(f"  Matrix: {X.shape}")

svd = TruncatedSVD(n_components=100, random_state=42)
X_lsa = svd.fit_transform(X)
X_lsa = normalize(X_lsa)
var_exp = svd.explained_variance_ratio_.sum()
print(f"  LSA variance explained (100 components): {var_exp:.2%}")

# ── SILHOUETTE SWEEP ─────────────────────────────────────────────────────────
print("Sweeping K=6..17 for optimal clusters...")
sil_scores = {}
for k in range(6, 18):
    km = KMeans(n_clusters=k, random_state=42, n_init=10, max_iter=300)
    labels = km.fit_predict(X_lsa)
    sil = silhouette_score(X_lsa, labels, sample_size=600, random_state=42)
    sil_scores[k] = sil
    print(f"  K={k:2d}: sil={sil:.4f}")

optimal_k = max(sil_scores, key=sil_scores.get)
print(f"\nOptimal K = {optimal_k}  (silhouette={sil_scores[optimal_k]:.4f})")

# ── FINAL K-MEANS ────────────────────────────────────────────────────────────
km_final = KMeans(n_clusters=optimal_k, random_state=42, n_init=20, max_iter=500)
df['cluster_id'] = km_final.fit_predict(X_lsa)

# ── UMAP 2D ──────────────────────────────────────────────────────────────────
print("Computing UMAP 2D projection...")
try:
    import umap as umap_lib
    reducer = umap_lib.UMAP(n_components=2, random_state=42, n_neighbors=15,
                            min_dist=0.1, metric='cosine', verbose=False)
    coords = reducer.fit_transform(X_lsa)
    df['umap_x'] = coords[:, 0].round(4)
    df['umap_y'] = coords[:, 1].round(4)
    print("  UMAP done.")
except Exception as e:
    print(f"  UMAP skipped: {e}")
    df['umap_x'] = 0.0
    df['umap_y'] = 0.0

# ── CLUSTER METADATA ─────────────────────────────────────────────────────────
feature_names = vectorizer.get_feature_names_out()
cluster_meta = {}

for cid in sorted(df['cluster_id'].unique()):
    mask = df['cluster_id'] == cid
    cr = df[mask]
    n = int(mask.sum())

    centroid = km_final.cluster_centers_[cid]
    centroid_tfidf = svd.inverse_transform(centroid.reshape(1, -1))[0]
    top_idx = centroid_tfidf.argsort()[::-1][:20]
    top_terms = [feature_names[i] for i in top_idx
                 if feature_names[i] not in CLINICAL_STOPWORDS and len(feature_names[i]) > 4][:10]

    arch_counts = cr['rule_archetype'].value_counts()
    dom_arch = arch_counts.index[0]
    arch_pct = arch_counts.iloc[0] / n * 100

    bs_counts = cr['BodySystem'].value_counts().head(3)
    top_bs = ', '.join([f"{v}({c})" for v, c in bs_counts.items()])

    yr_dist = {int(k): int(v) for k, v in cr['ExamYear'].value_counts().sort_index().items()}
    peak_yr = max(yr_dist, key=yr_dist.get)

    # Auto display name: archetype label + top clinical term
    arch_label = ARCHETYPE_DISPLAY.get(dom_arch, dom_arch)
    clinical_term = top_terms[0].replace('_',' ').title() if top_terms else ''
    display_name = f"{arch_label}: {clinical_term}" if clinical_term else arch_label

    cluster_meta[cid] = {
        'cluster_id': cid,
        'display_name': display_name,
        'n_questions': n,
        'pct_total': round(n / len(df) * 100, 1),
        'dominant_archetype': dom_arch,
        'archetype_pct': round(arch_pct, 1),
        'top_terms': top_terms,
        'top_body_systems': top_bs,
        'year_dist': yr_dist,
        'peak_year': peak_yr,
    }
    print(f"  C{cid:02d} [{n:3d}q] {display_name.encode('ascii','replace').decode()}")

df['cluster_display'] = df['cluster_id'].map({c: m['display_name'] for c, m in cluster_meta.items()})
df['cluster_short']   = df['cluster_id'].map({c: f"C{c:02d}" for c in cluster_meta})

# ── EXCEL BUILD ──────────────────────────────────────────────────────────────
CLUSTER_COLORS = [
    'D6E4F0','D5F5E3','FAD7A0','F5CBA7','D7BDE2',
    'AED6F1','A9DFBF','F9E79F','FDEBD0','E8DAEF',
    'ABEBC6','F1948A','A8D8EA','F7DC6F','FAE5D3','FEF9E7',
]
HDR_COLOR = '1F3864'
ACCENT    = '2E75B6'
SUBHDR    = 'BDD7EE'

def hdr(cell, bg=HDR_COLOR, fg='FFFFFF', bold=True, sz=10):
    cell.font  = Font(bold=bold, color=fg, name='Arial', size=sz)
    cell.fill  = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def tb():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def cw(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

wb = Workbook()

# ── TAB 1: CLUSTER SUMMARY ───────────────────────────────────────────────────
ws1 = wb.active
ws1.title = 'Cluster Summary'
ws1.sheet_view.showGridLines = False

ws1.merge_cells('A1:J1')
ws1['A1'] = f'ITE EXAM FINGERPRINT  |  1,200 Questions  |  2020-2025  |  K={optimal_k} Clusters'
ws1['A1'].font = Font(bold=True, size=14, color='FFFFFF', name='Arial')
ws1['A1'].fill = PatternFill('solid', start_color=HDR_COLOR)
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 32

ws1.merge_cells('A2:J2')
ws1['A2'] = (f'Method: TF-IDF (2500 features, bigrams) → LSA (100 components, {var_exp:.0%} variance) '
             f'→ K-Means  |  Best Silhouette: {sil_scores[optimal_k]:.4f}  |  Rule-based archetype overlay')
ws1['A2'].font = Font(italic=True, size=9, color='555555', name='Arial')
ws1['A2'].alignment = Alignment(horizontal='center')
ws1.row_dimensions[2].height = 14

cols_s = ['Cluster ID','Display Name','N Questions','% Total',
          'Dominant Archetype','Archetype %','Top Body Systems',
          'Top Clinical Terms','Peak Year','Year Trend (2020→2025)']
for c, h in enumerate(cols_s, 1):
    cell = ws1.cell(row=4, column=c, value=h)
    hdr(cell)
ws1.row_dimensions[4].height = 26

for ri, (cid, meta) in enumerate(sorted(cluster_meta.items()), 5):
    col = CLUSTER_COLORS[cid % len(CLUSTER_COLORS)]
    trend = '→'.join([str(meta['year_dist'].get(y, 0)) for y in range(2020, 2026)])
    vals = [f"C{cid:02d}", meta['display_name'], meta['n_questions'],
            f"{meta['pct_total']:.1f}%", ARCHETYPE_DISPLAY.get(meta['dominant_archetype'], meta['dominant_archetype']),
            f"{meta['archetype_pct']:.0f}%", meta['top_body_systems'],
            ', '.join(meta['top_terms'][:6]), str(meta['peak_year']), trend]
    for c, v in enumerate(vals, 1):
        cell = ws1.cell(row=ri, column=c, value=v)
        cell.fill = PatternFill('solid', start_color=col)
        cell.font = Font(name='Arial', size=9)
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = tb()
    ws1.row_dimensions[ri].height = 38

cw(ws1, {'A':10,'B':34,'C':12,'D':9,'E':20,'F':12,'G':30,'H':42,'I':10,'J':30})

# ── TAB 2: TAGGED QUESTIONS ──────────────────────────────────────────────────
ws2 = wb.create_sheet('Tagged Questions')
ws2.sheet_view.showGridLines = False

ws2.merge_cells('A1:I1')
ws2['A1'] = 'ALL 1,200 ITE QUESTIONS — CLUSTER & ARCHETYPE TAGGED'
ws2['A1'].font = Font(bold=True, size=13, color='FFFFFF', name='Arial')
ws2['A1'].fill = PatternFill('solid', start_color=HDR_COLOR)
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 26

q_hdrs = ['QuestionID','Year','Cluster','Archetype','Body System',
          'Subcategory','Blueprint','Question Stem (first 200 chars)','UMAP X','UMAP Y']
for c, h in enumerate(q_hdrs, 1):
    cell = ws2.cell(row=2, column=c, value=h)
    hdr(cell)
ws2.row_dimensions[2].height = 22
ws2.auto_filter.ref = 'A2:J2'
ws2.freeze_panes = 'A3'

for ri, (_, row) in enumerate(df.iterrows(), 3):
    cid = int(row['cluster_id'])
    col = CLUSTER_COLORS[cid % len(CLUSTER_COLORS)]
    stem_s = str(row['QuestionStem'])[:200].replace('\n',' ')
    vals = [row['QuestionID'], row['ExamYear'], row['cluster_display'],
            ARCHETYPE_DISPLAY.get(row['rule_archetype'], row['rule_archetype']),
            row['BodySystem'], str(row.get('Subcategory','')),
            str(row.get('BlueprintCategory_Predicted','')),
            stem_s, row['umap_x'], row['umap_y']]
    for c, v in enumerate(vals, 1):
        cell = ws2.cell(row=ri, column=c, value=v)
        cell.fill  = PatternFill('solid', start_color=col)
        cell.font  = Font(name='Arial', size=8)
        cell.alignment = Alignment(vertical='top', wrap_text=(c == 8))
        cell.border = tb()
    ws2.row_dimensions[ri].height = 38

cw(ws2, {'A':14,'B':6,'C':32,'D':18,'E':22,'F':18,'G':20,'H':62,'I':9,'J':9})

# ── TAB 3: CLUSTER × BODY SYSTEM HEATMAP ────────────────────────────────────
ws3 = wb.create_sheet('Cluster x BodySystem')
ws3.sheet_view.showGridLines = False

ws3.merge_cells('A1:R1')
ws3['A1'] = 'CLUSTER × BODY SYSTEM HEATMAP  (cell value = question count)'
ws3['A1'].font  = Font(bold=True, size=12, color='FFFFFF', name='Arial')
ws3['A1'].fill  = PatternFill('solid', start_color=HDR_COLOR)
ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 26

pivot = pd.crosstab(df['cluster_display'], df['BodySystem'])
body_systems = list(pivot.columns)
clusters_list = list(pivot.index)

ws3.cell(row=2, column=1, value='Cluster').font = Font(bold=True, name='Arial', size=10)
for c, bs in enumerate(body_systems, 2):
    cell = ws3.cell(row=2, column=c, value=bs)
    cell.font  = Font(bold=True, name='Arial', size=9, color='1F3864')
    cell.fill  = PatternFill('solid', start_color=SUBHDR)
    cell.alignment = Alignment(text_rotation=45, horizontal='center', vertical='bottom')
ws3.row_dimensions[2].height = 60

for ri, cl in enumerate(clusters_list, 3):
    ws3.cell(row=ri, column=1, value=cl).font = Font(name='Arial', size=9)
    row_max = pivot.loc[cl].max()
    for c, bs in enumerate(body_systems, 2):
        val = int(pivot.loc[cl, bs]) if bs in pivot.columns else 0
        cell = ws3.cell(row=ri, column=c, value=val if val > 0 else '')
        intensity = int(val / max(row_max, 1) * 190) if val > 0 else 0
        r_val = 255 - intensity
        gb_val = max(255 - intensity // 2, 40)
        hex_c = f'{r_val:02X}{r_val:02X}{gb_val:02X}'
        cell.fill  = PatternFill('solid', start_color=hex_c)
        cell.font  = Font(name='Arial', size=9, bold=(val == row_max and val > 0))
        cell.alignment = Alignment(horizontal='center')
        cell.border = tb()
    ws3.row_dimensions[ri].height = 20

ws3.column_dimensions['A'].width = 36
for c in range(2, len(body_systems) + 2):
    ws3.column_dimensions[get_column_letter(c)].width = 13

# ── TAB 4: ARCHETYPE × YEAR TREND ───────────────────────────────────────────
ws4 = wb.create_sheet('Archetype x Year')
ws4.sheet_view.showGridLines = False

ws4.merge_cells('A1:J1')
ws4['A1'] = 'QUESTION ARCHETYPE DISTRIBUTION BY EXAM YEAR  (2020–2025)'
ws4['A1'].font  = Font(bold=True, size=12, color='FFFFFF', name='Arial')
ws4['A1'].fill  = PatternFill('solid', start_color=HDR_COLOR)
ws4['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws4.row_dimensions[1].height = 26

arch_year = pd.crosstab(df['rule_archetype'], df['ExamYear'])
years = sorted(df['ExamYear'].unique())
archetypes = list(arch_year.index)

ws4.cell(row=2, column=1, value='Archetype').font = Font(bold=True, name='Arial', size=10)
for c, yr in enumerate(years, 2):
    cell = ws4.cell(row=2, column=c, value=str(yr))
    hdr(cell, sz=10)
hdr(ws4.cell(row=2, column=len(years)+2, value='TOTAL'), bg=ACCENT, sz=10)
hdr(ws4.cell(row=2, column=len(years)+3, value='TREND'), bg=ACCENT, sz=10)
ws4.row_dimensions[2].height = 22

ARCH_COLORS = {
    'DIAGNOSIS':'D6EAF8','NEXT_STEP':'D5F5E3','MANAGEMENT':'FDEBD0',
    'SCREENING':'FEF9E7','COUNSELING':'F4ECF7','PHARMACOLOGY':'EAFAF1',
    'INTERPRETATION':'EBF5FB','REFERRAL':'FDEDEC','SURVEILLANCE':'FFF9E6',
    'RISK_FACTORS':'F8F9FA','UNCLASSIFIED':'F2F3F4',
}

for ri, arch in enumerate(archetypes, 3):
    col = ARCH_COLORS.get(arch, 'FFFFFF')
    label = ARCHETYPE_DISPLAY.get(arch, arch)
    cell = ws4.cell(row=ri, column=1, value=label)
    cell.fill = PatternFill('solid', start_color=col)
    cell.font = Font(name='Arial', size=10, bold=True)
    year_vals = []
    for c, yr in enumerate(years, 2):
        val = int(arch_year.loc[arch, yr]) if yr in arch_year.columns else 0
        year_vals.append(val)
        cell = ws4.cell(row=ri, column=c, value=val)
        cell.fill = PatternFill('solid', start_color=col)
        cell.font = Font(name='Arial', size=10)
        cell.alignment = Alignment(horizontal='center')
        cell.border = tb()
    total = sum(year_vals)
    tc = ws4.cell(row=ri, column=len(years)+2, value=total)
    tc.font = Font(bold=True, name='Arial', size=10)
    if year_vals and year_vals[0] > 0:
        delta = year_vals[-1] - year_vals[0]
        sym = '↑' if delta > 2 else ('↓' if delta < -2 else '→')
    else:
        sym = '—'
    tc2 = ws4.cell(row=ri, column=len(years)+3, value=sym)
    tc2.font = Font(name='Arial', size=14, bold=True,
                    color=('1E8449' if sym=='↑' else ('C0392B' if sym=='↓' else '2C3E50')))
    tc2.alignment = Alignment(horizontal='center')
    ws4.row_dimensions[ri].height = 22

cw(ws4, {'A':26,'B':8,'C':8,'D':8,'E':8,'F':8,'G':8,'H':10,'I':10})

# ── TAB 5: METHODOLOGY ───────────────────────────────────────────────────────
ws5 = wb.create_sheet('Method Notes')
ws5.sheet_view.showGridLines = False

ws5.merge_cells('A1:D1')
ws5['A1'] = 'CLUSTERING METHODOLOGY & SILHOUETTE SWEEP'
ws5['A1'].font  = Font(bold=True, size=12, color='FFFFFF', name='Arial')
ws5['A1'].fill  = PatternFill('solid', start_color=HDR_COLOR)
ws5['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws5.row_dimensions[1].height = 26

ws5.cell(row=3, column=1, value='K').font = Font(bold=True, name='Arial')
ws5.cell(row=3, column=2, value='Silhouette Score').font = Font(bold=True, name='Arial')
ws5.cell(row=3, column=3, value='Selected?').font = Font(bold=True, name='Arial')

for ri, (k, sil) in enumerate(sorted(sil_scores.items()), 4):
    ws5.cell(row=ri, column=1, value=k)
    ws5.cell(row=ri, column=2, value=round(sil, 4))
    is_opt = (k == optimal_k)
    ws5.cell(row=ri, column=3, value='✓ OPTIMAL' if is_opt else '')
    if is_opt:
        for c in range(1, 4):
            ws5.cell(row=ri, column=c).fill = PatternFill('solid', start_color='D5F5E3')
            ws5.cell(row=ri, column=c).font = Font(bold=True, name='Arial', color='1E8449')

notes = [
    '', 'METHOD NOTES:',
    '• TF-IDF: 2,500 features, unigram+bigram, sublinear_tf=True, min_df=3, max_df=0.85',
    f'• LSA: 100 components, {var_exp:.1%} variance explained, cosine-normalized',
    '• K-Means: 20 initializations, 500 max iterations on normalized LSA vectors',
    '• UMAP: 2D projection, n_neighbors=15, min_dist=0.1, cosine metric',
    '• Archetype detection: 10 categories via regex on full question stem',
    '• Clinical stopwords: ~60 noise terms removed (patient, year, history, etc.)',
    '', 'ARCHETYPE DEFINITIONS:',
    '🔬 DIAGNOSIS — identify condition from clinical scenario',
    '⚡ NEXT_STEP — what to do immediately in acute scenario',
    '💊 MANAGEMENT — chronic disease treatment plans',
    '🛡️ SCREENING — preventive care, USPSTF guideline application',
    '💬 COUNSELING — what to tell the patient',
    '💉 PHARMACOLOGY — drug selection, dosing, interactions',
    '📊 INTERPRETATION — lab/imaging/ECG/biopsy findings',
    '🏥 REFERRAL/ADMIT — when to escalate to specialist or ER',
    '🔄 SURVEILLANCE — follow-up intervals, monitoring',
    '⚠️ RISK_FACTORS — etiology, predisposing conditions',
]
start_row = 4 + len(sil_scores) + 2
for i, note in enumerate(notes):
    cell = ws5.cell(row=start_row+i, column=1, value=note)
    if note.endswith(':') and note.strip():
        cell.font = Font(bold=True, name='Arial', size=10)
    else:
        cell.font = Font(name='Arial', size=9)

cw(ws5, {'A':62,'B':18,'C':14})

# ── SAVE ─────────────────────────────────────────────────────────────────────
print(f"\nSaving XLSX to: {OUT_XLSX}")
wb.save(OUT_XLSX)
print("  XLSX saved.")

# ── JSON SUMMARY ─────────────────────────────────────────────────────────────
with open(OUT_JSON, 'w', encoding='utf-8') as f:
    json.dump({str(k): v for k, v in cluster_meta.items()}, f, indent=2, default=str)
print(f"  JSON saved: {OUT_JSON}")

print("\n=== CLUSTER OVERVIEW ===")
for cid, meta in sorted(cluster_meta.items()):
    name_safe = meta['display_name'].encode('ascii','replace').decode()
    print(f"  C{cid:02d} [{meta['n_questions']:3d}q | {meta['pct_total']:4.1f}%] "
          f"{name_safe}  |  top: {', '.join(meta['top_terms'][:4])}")

print("\nDone.")
